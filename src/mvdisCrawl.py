#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# VERSION: PAGINATION_SCOPED_LINK_SERVER_CONFIRMED_2026-08-01
"""
監理服務網法人交通違規雙階段批次查詢工具。

主要流程：
1. 從 Excel 讀取統一編號／登記編號與登記名稱。
2. 使用 Playwright 先查詢「交通違規（含強制險）查詢及繳納」。
3. 分別完整擷取「可線上繳納」與「不可線上繳納」的每一頁。
4. 再查詢「交通違規繳納記錄查詢」，逐頁擷取所有已繳納紀錄。
5. CAPTCHA 圖片以 PIL.Image.Image 傳給：

       from englishAlphanumericOcrApi import ocrImage
       code = ocrImage(image)

6. 依網站 showbanner 驗證：
   - 實際走訪頁碼等於網站宣告總頁數。
   - 擷取筆數等於網站宣告總筆數。
   - 唯一資料筆數等於網站宣告總筆數。
   - 不得有同頁或跨頁重複擷取。
7. 完整性不符時立即重爬同一家公司，不會直接處理下一家公司；
   達重試上限後預設停止整批，避免把不完整結果標記為成功。
8. Excel 主表新增三組分頁資訊、完整性檢查與重複資料檢查欄位。
9. 公司明細工作表 A:G 保留原 155598 格式，H:N 註明每筆資料來源頁碼、
   總頁數、網站宣告總筆數、分頁文字與資料唯一鍵。
10. 新增「分頁完整性稽核」及「重複資料稽核」工作表。
11. 所有未解決錯誤與分頁驗證失敗均寫入 results\\mvdis_errorLog 的 JSON。
12. 永遠寫入輸出 Excel，支援逐筆儲存與中斷續跑；舊版未經分頁驗證的
    「成功」資料會自動重新查詢。

僅可查詢您有權處理的法人／商業資料，並請遵守監理服務網使用規範。
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict, deque
from copy import copy
import hashlib
from html import unescape
import json
import os
import random
import re
import shutil
import tempfile
import time
import traceback
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence
from urllib.parse import (
    parse_qsl,
    urlencode,
    urljoin,
    urlsplit,
    urlunsplit,
)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "缺少 openpyxl。請先執行：\n"
        "python -m pip install openpyxl playwright pillow\n"
        "python -m playwright install chromium"
    ) from exc

try:
    from PIL import Image
except ImportError as exc:
    raise SystemExit(
        "缺少 Pillow。請先執行：python -m pip install pillow"
    ) from exc

try:
    from englishAlphanumericOcrApi import ocrImage
except ImportError as exc:
    raise SystemExit(
        "無法從 englishAlphanumericOcrApi.py 匯入 ocrImage。\n"
        "請將 englishAlphanumericOcrApi.py 放在本腳本同一資料夾，\n"
        "並確認其中有 def ocrImage(image: Image.Image) -> str"
    ) from exc

try:
    from playwright.sync_api import (
        Browser,
        Error as PlaywrightError,
        Locator,
        Page,
        TimeoutError as PlaywrightTimeoutError,
        sync_playwright,
    )
except ImportError as exc:
    raise SystemExit(
        "缺少 Playwright。請先執行：\n"
        "python -m pip install playwright\n"
        "python -m playwright install chromium"
    ) from exc


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

# mvdisCrawl.py 位於 src 資料夾，因此從專案根目錄讀取：
# Data\高雄市.xlsx
DEFAULT_INPUT_PATH = PROJECT_ROOT / "Data" / "高雄市.xlsx"

# 輸出資料夾放在專案根目錄下的 results。
DEFAULT_RESULTS_DIR = PROJECT_ROOT / "results"
DEFAULT_CAPTCHA_DIR = DEFAULT_RESULTS_DIR / "mvdis_captcha"
DEFAULT_DEBUG_DIR = DEFAULT_RESULTS_DIR / "mvdis_debug"
DEFAULT_ERROR_LOG_DIR = DEFAULT_RESULTS_DIR / "mvdis_errorLog"

DEFAULT_URL = (
    "https://www.mvdis.gov.tw/m3-emv-vil/vil/penaltyQueryPayRecord/legal"
)

CAPTCHA_CAPTURE_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/div[2]/div/div/form/"
    "table/tbody/tr[4]/td"
)

# CAPTCHA 本身的精確圖片 XPath。優先截取 img，避免把整個 td 的
# 標籤、輸入框及「換一張」文字一起送進 OCR。
CAPTCHA_IMAGE_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/div[2]/div/div/form/"
    "table/tbody/tr[4]/td/img"
)

# 監理服務網法人違規查詢頁面的實際查詢按鈕是 <a> 元素。
QUERY_BUTTON_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/div[2]/div/div/form/div/a"
)

# 查詢結果頁中，完整「罰鍰繳納紀錄」JSON 所在的 hidden input。
# 使用者提供的實際 XPath：
RESULT_JSON_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/div[3]/form/input"
)

# 查詢結果頁顯示「查無已繳納罰鍰資料」的精確 XPath。
# 這個節點出現時，代表網站已正常完成查詢，只是結果筆數為 0，
# 不應寫成「錯誤」或「無資料」，而應寫成「成功」與 0 筆。
NO_PAID_DATA_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/div[1]/table/"
    "tbody/tr/td[1]"
)

# 第一階段：「交通違規（含強制險）查詢及繳納」。
# 依使用者指定，先從目前的法人繳費紀錄入口頁進入；若該入口頁
# 沒有載入新版查詢頁，則使用監理服務網正式的 penaltyQueryPay 路徑備援。
DEFAULT_CURRENT_QUERY_URL = DEFAULT_URL
CURRENT_QUERY_DIRECT_FALLBACK_URL = (
    "https://www.mvdis.gov.tw/m3-emv-vil/vil/penaltyQueryPay"
)

# 「交通違規（含強制險）查詢及繳納」法人頁籤。
CURRENT_QUERY_CORPORATE_TAB_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/div[3]/table/"
    "tbody/tr/td[2]/span/img"
)

# 法人資料與 CAPTCHA 填完後的查詢按鈕。
CURRENT_QUERY_BUTTON_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/div[3]/div/"
    "div[2]/form/div/a"
)

# 第一階段查詢結果容器與目前頁籤所顯示的表格。
CURRENT_RESULT_CONTAINER_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/form/div[2]/div[2]/div"
)
CURRENT_RESULT_TABLE_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/form/div[2]/div[2]/"
    "div/div[2]/table"
)

# 結果頁第二個頁籤；點擊後 CURRENT_RESULT_TABLE_XPATH 會切換成
# 另一組（不可線上繳納／需臨櫃處理）資料。
CURRENT_RESULT_SECOND_TAB_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/form/div[2]/"
    "table[2]/tbody/tr/td[2]/span/img"
)

# 第二階段「交通違規繳納記錄查詢」分頁資訊。
# 範例：1 / 27 頁，共 262 筆資料。
PAID_PAGINATION_BANNER_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/div[3]/form/div[1]/span"
)

# 第一階段「交通違規（含強制險）查詢及繳納」分頁資訊。
# 可線上繳納與不可線上繳納頁籤共用此位置。
CURRENT_PAGINATION_BANNER_XPATH = (
    "/html/body/table/tbody/tr[2]/td[1]/form/div[2]/div[2]/"
    "div/div[2]/div/span"
)

QUERY_SECTION_CURRENT = "交通違規（含強制險）查詢及繳納"
QUERY_SECTION_PAID = "交通違規繳納記錄查詢"
PAID_RECORD_CATEGORY = "已繳納紀錄"

PAGINATION_VERIFIED_MARKER = "分頁完整性已驗證"
PAGINATION_AUDIT_SHEET_NAME = "分頁完整性稽核"
DUPLICATE_AUDIT_SHEET_NAME = "重複資料稽核"
DETAIL_AUDIT_FIRST_COLUMN = 8
DETAIL_AUDIT_LAST_COLUMN = 14

COMBINED_QUERY_MESSAGE_MARKER = "雙階段查詢完成"
CURRENT_ONLINE_CATEGORY = "可線上繳納"
CURRENT_OFFLINE_CATEGORY = "不可線上繳納"

DETAIL_TEMPLATE_SHEET_NAME = "__違規明細範本__"
DEFAULT_DETAIL_TEMPLATE_SHEET = "155598"

TABLE_KIND_UNPAID = "unpaid"
TABLE_KIND_PAID = "paid"

LEGACY_BORDER_COLOR = "CACACA"
LEGACY_PAID_FONT_COLOR = "4D4D4D"
LEGACY_UNPAID_FONT_COLOR = "FF0000"
LEGACY_NEEDS_APPEARANCE_FONT_COLOR = "1F4E78"
LEGACY_ALT_FILL_COLOR = "F5F5F5"

STATUS_SUCCESS = "成功"
STATUS_NO_DATA = "無資料"
STATUS_ERROR = "錯誤"
STATUS_RETRY = "待重試"
STATUS_SKIPPED = "略過"

HEADER_ID_PATTERNS = ("統一編號", "登記編號")
HEADER_NAME_PATTERNS = ("登記名稱", "公司名稱", "商業名稱", "名稱")

RESULT_KEYWORDS = (
    "違規日期",
    "違規事由",
    "罰鍰",
    "舉發單號",
    "車牌",
    "應到案",
    "繳款",
    "處理方式",
)

NO_DATA_PATTERNS = (
    "查無資料",
    "查無違規",
    "無違規資料",
    "目前無違規",
    "沒有違規資料",
    "無符合資料",
    "查無繳費紀錄",
    # 結果頁實際使用的訊息。舊版遺漏這句，導致正常無資料
    # 被錯誤寫成「錯誤」。
    "查無已繳納罰鍰資料",
    "無已繳納罰鍰資料",
    "查無罰鍰繳納資料",
    "查無可線上繳納罰單資料",
    "查無不可線上繳納罰單資料",
    "查無交通違規資料",
    "無交通違規資料",
)

CAPTCHA_ERROR_PATTERNS = (
    "驗證碼錯誤",
    "驗證碼不正確",
    "驗證碼有誤",
    "請輸入驗證碼",
    "驗證碼輸入錯誤",
    "驗證碼輸入有誤",
    "驗證碼驗證失敗",
)

FORM_ERROR_PATTERNS = (
    "請輸入統一編號",
    "統一編號錯誤",
    "統一編號格式",
    "系統忙碌",
    "查詢失敗",
    "操作逾時",
)

# 這些不是永久資料錯誤，應重新查詢，而不是直接寫成「錯誤」。
TRANSIENT_FORM_ERROR_PATTERNS = (
    "系統忙碌",
    "查詢失敗",
    "操作逾時",
)

THIN_GRAY = Side(style="thin", color="D9E2F3")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
SUBTITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD_FONT = Font(bold=True)


@dataclass(frozen=True)
class CompanyRow:
    excel_row: int
    raw_id: str
    query_id: str
    name: str


@dataclass(frozen=True)
class PaginationInfo:
    section: str
    category: str
    current_page: int
    total_pages: int
    declared_total_records: int
    banner_text: str
    raw_text: str


@dataclass
class PageAudit:
    section: str
    category: str
    page_number: int
    total_pages: int
    declared_total_records: int
    banner_text: str
    extracted_count: int
    record_keys: list[str]
    result_url: str

    @property
    def unique_count(self) -> int:
        return len(set(self.record_keys))

    @property
    def duplicate_count(self) -> int:
        return max(0, len(self.record_keys) - self.unique_count)


@dataclass
class ExtractedTable:
    kind: str
    title: str
    headers: list[str]
    rows: list[list[Any]]
    row_keys: list[str] = field(default_factory=list)
    section: str = ""
    category: str = ""
    page_number: int = 1
    total_pages: int = 1
    declared_total_records: int | None = None
    banner_text: str = ""

    @property
    def record_count(self) -> int:
        return len(self.rows)


@dataclass(frozen=True)
class DetailRecord:
    kind: str
    row: list[Any]
    section: str
    category: str
    page_number: int
    total_pages: int
    declared_total_records: int
    banner_text: str
    record_key: str


@dataclass
class QueryOutcome:
    status: str
    record_count: int
    tables: list[ExtractedTable]
    message: str
    result_url: str
    retry_later: bool = False
    page_audits: list[PageAudit] = field(default_factory=list)
    duplicate_keys: list[str] = field(default_factory=list)


class CaptchaRecognitionError(RuntimeError):
    """OCR 沒有產生符合格式的四碼字串；可重新取得 CAPTCHA。"""


class TransientPageError(RuntimeError):
    """頁面仍在導向或結果尚未完整載入；應重新查詢。"""


class DataIntegrityError(RuntimeError):
    """分頁、筆數或重複資料驗證失敗；不得直接處理下一家公司。"""

    def __init__(
        self,
        message: str,
        details: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.details = details or {}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_company_id(value: Any) -> tuple[str, str]:
    """Return (raw display id, 8-digit query id)."""
    if value is None:
        raise ValueError("統一編號是空白")

    if isinstance(value, bool):
        raise ValueError("統一編號格式錯誤")

    if isinstance(value, int):
        raw = str(value)
    elif isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"統一編號不是整數：{value}")
        raw = str(int(value))
    else:
        raw = normalize_text(value)
        if raw.endswith(".0") and raw[:-2].isdigit():
            raw = raw[:-2]

    digits = re.sub(r"\D", "", raw)

    if not digits:
        raise ValueError(f"找不到數字：{raw!r}")

    if len(digits) > 8:
        raise ValueError(f"統一編號超過 8 碼：{raw!r}")

    return raw, digits.zfill(8)


def find_main_sheet_and_columns(
    workbook: Any,
) -> tuple[Any, int, int, int, int]:
    """Return worksheet, header_start_row, data_start_row, id_col, name_col."""
    for worksheet in workbook.worksheets:
        max_scan_row = min(max(worksheet.max_row, 1), 12)
        max_scan_col = min(max(worksheet.max_column, 1), 30)

        id_candidates: list[tuple[int, int]] = []
        name_candidates: list[tuple[int, int]] = []

        for row in range(1, max_scan_row + 1):
            for col in range(1, max_scan_col + 1):
                text = normalize_text(worksheet.cell(row, col).value)

                if not text:
                    continue

                if any(pattern in text for pattern in HEADER_ID_PATTERNS):
                    id_candidates.append((row, col))

                if any(pattern in text for pattern in HEADER_NAME_PATTERNS):
                    name_candidates.append((row, col))

        if not id_candidates or not name_candidates:
            continue

        id_row, id_col = min(id_candidates)

        compatible_names = [
            item
            for item in name_candidates
            if abs(item[0] - id_row) <= 2
        ]

        name_row, name_col = min(
            compatible_names or name_candidates
        )

        header_start = min(id_row, name_row)
        header_end = max(id_row, name_row)

        for probe_row in range(
            header_start,
            min(header_start + 3, worksheet.max_row) + 1,
        ):
            row_text = " ".join(
                normalize_text(worksheet.cell(probe_row, col).value)
                for col in range(1, max_scan_col + 1)
            )

            if any(
                pattern in row_text
                for pattern in HEADER_ID_PATTERNS
            ):
                header_end = max(header_end, probe_row)

        return (
            worksheet,
            header_start,
            header_end + 1,
            id_col,
            name_col,
        )

    raise ValueError(
        "找不到同時包含『統一編號／登記編號』與『登記名稱』的主工作表"
    )


def iter_companies(
    worksheet: Any,
    data_start_row: int,
    id_col: int,
    name_col: int,
    start_row: int | None,
    end_row: int | None,
    limit: int | None,
) -> list[CompanyRow]:
    first = max(data_start_row, start_row or data_start_row)
    last = min(worksheet.max_row, end_row or worksheet.max_row)

    companies: list[CompanyRow] = []

    for row in range(first, last + 1):
        value = worksheet.cell(row, id_col).value
        name = normalize_text(worksheet.cell(row, name_col).value)

        if value in (None, ""):
            continue

        try:
            raw_id, query_id = normalize_company_id(value)
        except ValueError:
            continue

        companies.append(
            CompanyRow(
                excel_row=row,
                raw_id=raw_id,
                query_id=query_id,
                name=name,
            )
        )

        if limit is not None and len(companies) >= limit:
            break

    return companies


def ensure_tracking_columns(
    worksheet: Any,
    header_start_row: int,
) -> dict[str, int]:
    """建立主表追蹤欄位，包含三組分頁資訊與完整性稽核結果。"""
    desired = [
        "交通違規筆數",
        "違規查詢狀態",
        "最後查詢時間",
        "違規查詢訊息",
        "可線上繳納分頁資訊",
        "不可線上繳納分頁資訊",
        "繳納紀錄分頁資訊",
        "分頁完整性檢查",
        "重複資料檢查",
    ]

    existing: dict[str, int] = {}

    for col in range(1, worksheet.max_column + 1):
        text = normalize_text(
            worksheet.cell(header_start_row, col).value
        )

        if text in desired:
            existing[text] = col

    if "交通違規筆數" not in existing:
        numeric_blank_candidates: list[tuple[int, int]] = []

        for col in range(1, worksheet.max_column + 1):
            if normalize_text(
                worksheet.cell(header_start_row, col).value
            ):
                continue

            numeric_count = 0

            for row in range(
                header_start_row + 1,
                min(
                    worksheet.max_row,
                    header_start_row + 250,
                ) + 1,
            ):
                value = worksheet.cell(row, col).value

                if (
                    isinstance(value, (int, float))
                    and not isinstance(value, bool)
                ):
                    numeric_count += 1

            if numeric_count:
                numeric_blank_candidates.append(
                    (col, numeric_count)
                )

        if numeric_blank_candidates:
            existing["交通違規筆數"] = max(
                numeric_blank_candidates
            )[0]

    next_col = worksheet.max_column + 1

    for header in desired:
        if header not in existing:
            existing[header] = next_col
            worksheet.cell(
                header_start_row,
                next_col,
                header,
            )
            next_col += 1

    widths = {
        "交通違規筆數": 14,
        "違規查詢狀態": 14,
        "最後查詢時間": 20,
        "違規查詢訊息": 58,
        "可線上繳納分頁資訊": 40,
        "不可線上繳納分頁資訊": 40,
        "繳納紀錄分頁資訊": 40,
        "分頁完整性檢查": 26,
        "重複資料檢查": 24,
    }

    for header in desired:
        col = existing[header]
        cell = worksheet.cell(header_start_row, col)
        cell.value = header
        cell.fill = TITLE_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(
            left=THIN_GRAY,
            right=THIN_GRAY,
            top=THIN_GRAY,
            bottom=THIN_GRAY,
        )

        worksheet.column_dimensions[
            get_column_letter(col)
        ].width = widths[header]

    return existing

def reset_tracking_values(
    worksheet: Any,
    data_start_row: int,
    tracking_columns: dict[str, int],
) -> None:
    for row in range(
        data_start_row,
        worksheet.max_row + 1,
    ):
        for col in tracking_columns.values():
            worksheet.cell(row, col).value = None


def normalized_numeric_sheet_title(value: Any) -> str:
    text = normalize_text(value)
    digits = re.sub(r"\D", "", text)

    if not digits:
        return ""

    return digits.lstrip("0") or "0"


def detail_sheet_name(
    raw_id: str,
    query_id: str,
) -> str:
    """
    與來源檔一致，工作表名稱使用 Excel 中顯示的登記編號。

    例如：
        查詢用編號：00155598
        工作表名稱：155598
    """
    normalized_raw = normalized_numeric_sheet_title(
        raw_id
    )

    normalized_query = normalized_numeric_sheet_title(
        query_id
    )

    return safe_sheet_name(
        normalized_raw
        or normalized_query
        or query_id
    )


def clear_stale_numeric_detail_sheets(
    workbook: Any,
    main_title: str,
) -> None:
    for worksheet in list(workbook.worksheets):
        title = worksheet.title.strip()

        if title in {
            main_title,
            DETAIL_TEMPLATE_SHEET_NAME,
        }:
            continue

        if re.fullmatch(r"\d{1,8}", title):
            workbook.remove(worksheet)


def safe_sheet_name(company_id: str) -> str:
    return (
        re.sub(r"[\\/*?:\[\]]", "_", company_id)[:31]
        or "明細"
    )


def remove_existing_detail_variants(
    workbook: Any,
    raw_id: str,
    query_id: str,
) -> None:
    candidates = {
        safe_sheet_name(raw_id),
        safe_sheet_name(query_id),
        detail_sheet_name(
            raw_id,
            query_id,
        ),
    }

    candidates.discard(
        DETAIL_TEMPLATE_SHEET_NAME
    )

    for title in list(candidates):
        if title in workbook.sheetnames:
            workbook.remove(
                workbook[title]
            )


def legacy_medium_side() -> Side:
    return Side(
        style="medium",
        color=LEGACY_BORDER_COLOR,
    )


def legacy_border(
    *,
    left: bool,
    right: bool = True,
    top: bool,
    bottom: bool = True,
) -> Border:
    return Border(
        left=(
            legacy_medium_side()
            if left
            else Side()
        ),
        right=(
            legacy_medium_side()
            if right
            else Side()
        ),
        top=(
            legacy_medium_side()
            if top
            else Side()
        ),
        bottom=(
            legacy_medium_side()
            if bottom
            else Side()
        ),
    )


def apply_legacy_detail_row_style(
    worksheet: Any,
    row: int,
    *,
    column_count: int,
    font_color: str,
    first_row: bool,
    fill_color: str | None,
    hyperlink_column: int | None,
    row_height: float,
) -> None:
    """
    建立與來源檔 155598 工作表相同的基本視覺格式。

    - Arial
    - 一般內容 6 pt
    - 「檢視」10 pt、底線
    - #CACACA medium 框線
    - 垂直置中、文字換行
    - 交替灰底 #F5F5F5
    """
    for column in range(
        1,
        column_count + 1,
    ):
        cell = worksheet.cell(
            row,
            column,
        )

        is_hyperlink = (
            hyperlink_column == column
        )

        cell.font = Font(
            name="Arial",
            size=(
                10
                if is_hyperlink
                else 6
            ),
            color=font_color,
            underline=(
                "single"
                if is_hyperlink
                else None
            ),
        )

        cell.fill = (
            PatternFill(
                "solid",
                fgColor=fill_color,
            )
            if fill_color
            else PatternFill(
                fill_type=None
            )
        )

        cell.border = legacy_border(
            left=(column == 1),
            top=first_row,
        )

        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[
        row
    ].height = row_height


def build_fallback_detail_template(
    worksheet: Any,
) -> None:
    """
    當續跑檔案中不存在來源 155598 工作表時，
    以程式建立等價的隱藏範本。
    """
    worksheet.sheet_format.defaultRowHeight = 12.5
    worksheet.freeze_panes = None
    worksheet.sheet_view.showGridLines = True

    try:
        worksheet.sheet_properties.pageSetUpPr.fitToPage = False
    except Exception:
        pass

    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.paperSize = (
        worksheet.PAPERSIZE_A4
    )

    worksheet.page_margins.left = 0.7
    worksheet.page_margins.right = 0.7
    worksheet.page_margins.top = 0.75
    worksheet.page_margins.bottom = 0.75
    worksheet.page_margins.header = 0.3
    worksheet.page_margins.footer = 0.3

    # 155598 範本：
    # 1：第一筆一般未繳
    # 2～4：後續一般未繳
    # 5：需到案
    # 6：第一筆繳納紀錄
    # 7～10：中間繳納紀錄
    # 11：最後一筆繳納紀錄（灰底）
    apply_legacy_detail_row_style(
        worksheet,
        1,
        column_count=6,
        font_color=(
            LEGACY_UNPAID_FONT_COLOR
        ),
        first_row=True,
        fill_color=None,
        hyperlink_column=6,
        row_height=24.5,
    )

    for row in range(2, 5):
        apply_legacy_detail_row_style(
            worksheet,
            row,
            column_count=6,
            font_color=(
                LEGACY_UNPAID_FONT_COLOR
            ),
            first_row=False,
            fill_color=None,
            hyperlink_column=6,
            row_height=24.5,
        )

    apply_legacy_detail_row_style(
        worksheet,
        5,
        column_count=6,
        font_color=(
            LEGACY_NEEDS_APPEARANCE_FONT_COLOR
        ),
        first_row=True,
        fill_color=(
            LEGACY_ALT_FILL_COLOR
        ),
        hyperlink_column=6,
        row_height=16.5,
    )

    apply_legacy_detail_row_style(
        worksheet,
        6,
        column_count=7,
        font_color=(
            LEGACY_PAID_FONT_COLOR
        ),
        first_row=True,
        fill_color=None,
        hyperlink_column=None,
        row_height=16.5,
    )

    for row in range(7, 11):
        apply_legacy_detail_row_style(
            worksheet,
            row,
            column_count=7,
            font_color=(
                LEGACY_PAID_FONT_COLOR
            ),
            first_row=False,
            fill_color=None,
            hyperlink_column=None,
            row_height=16.5,
        )

    apply_legacy_detail_row_style(
        worksheet,
        11,
        column_count=7,
        font_color=(
            LEGACY_PAID_FONT_COLOR
        ),
        first_row=False,
        fill_color=(
            LEGACY_ALT_FILL_COLOR
        ),
        hyperlink_column=None,
        row_height=16.5,
    )



def is_legacy_detail_template_sheet(
    worksheet: Any,
) -> bool:
    if (
        worksheet.max_column > 7
        or worksheet.max_row < 1
    ):
        return False

    first_values = {
        normalize_text(
            worksheet.cell(
                row,
                column,
            ).value
        )
        for row in range(
            1,
            min(
                worksheet.max_row,
                8,
            ) + 1,
        )
        for column in range(
            1,
            min(
                worksheet.max_column,
                7,
            ) + 1,
        )
    }

    if first_values.intersection(
        {
            "統一編號",
            "登記名稱",
            "查詢狀態",
            "違規筆數",
            "查詢時間",
            "來源網址",
        }
    ):
        return False

    return True


def find_detail_template_candidate(
    workbook: Any,
    main_title: str,
    preferred_title: str,
) -> Any | None:
    preferred_normalized = normalize_text(
        preferred_title
    )

    if (
        preferred_normalized
        and preferred_normalized
        in workbook.sheetnames
    ):
        preferred_sheet = workbook[
            preferred_normalized
        ]

        if is_legacy_detail_template_sheet(
            preferred_sheet
        ):
            return preferred_sheet

    preferred_numeric = (
        normalized_numeric_sheet_title(
            preferred_normalized
        )
    )

    for worksheet in workbook.worksheets:
        title = normalize_text(
            worksheet.title
        )

        if title in {
            main_title,
            DETAIL_TEMPLATE_SHEET_NAME,
            "違規查詢紀錄",
        }:
            continue

        numeric_title = (
            normalized_numeric_sheet_title(
                title
            )
        )

        if (
            preferred_numeric
            and numeric_title
            == preferred_numeric
            and is_legacy_detail_template_sheet(
                worksheet
            )
        ):
            return worksheet

    candidates: list[Any] = []

    for worksheet in workbook.worksheets:
        title = normalize_text(
            worksheet.title
        )

        if title in {
            main_title,
            DETAIL_TEMPLATE_SHEET_NAME,
            "違規查詢紀錄",
        }:
            continue

        if not re.fullmatch(
            r"\d{1,8}",
            title,
        ):
            continue

        if is_legacy_detail_template_sheet(
            worksheet
        ):
            candidates.append(
                worksheet
            )

    if not candidates:
        return None

    return min(
        candidates,
        key=lambda item: (
            abs(item.max_row - 11),
            item.max_column,
        ),
    )



def copy_external_detail_template(
    source: Any,
    target: Any,
) -> None:
    """
    將來源活頁簿的 155598 格式複製到既有結果活頁簿。

    不直接複製 StyleArray 編號，改為逐項複製字型、底色、
    框線與對齊，避免跨活頁簿 style id 不一致。
    """
    for row in source.iter_rows():
        for source_cell in row:
            target_cell = target.cell(
                source_cell.row,
                source_cell.column,
            )

            target_cell.value = (
                source_cell.value
            )

            if source_cell.has_style:
                target_cell.font = copy(
                    source_cell.font
                )
                target_cell.fill = copy(
                    source_cell.fill
                )
                target_cell.border = copy(
                    source_cell.border
                )
                target_cell.alignment = copy(
                    source_cell.alignment
                )
                target_cell.number_format = (
                    source_cell.number_format
                )
                target_cell.protection = copy(
                    source_cell.protection
                )

            if source_cell.hyperlink:
                target_cell.hyperlink = copy(
                    source_cell.hyperlink
                )

    for row_index, dimension in (
        source.row_dimensions.items()
    ):
        target_dimension = (
            target.row_dimensions[
                row_index
            ]
        )

        target_dimension.height = (
            dimension.height
        )
        target_dimension.hidden = (
            dimension.hidden
        )
        target_dimension.outlineLevel = (
            dimension.outlineLevel
        )
        target_dimension.thickTop = getattr(
            dimension,
            "thickTop",
            False,
        )
        target_dimension.thickBot = getattr(
            dimension,
            "thickBot",
            False,
        )

    for column_letter, dimension in (
        source.column_dimensions.items()
    ):
        target_dimension = (
            target.column_dimensions[
                column_letter
            ]
        )

        target_dimension.width = (
            dimension.width
        )
        target_dimension.hidden = (
            dimension.hidden
        )
        target_dimension.bestFit = (
            dimension.bestFit
        )
        target_dimension.outlineLevel = (
            dimension.outlineLevel
        )

    target.sheet_format = copy(
        source.sheet_format
    )

    target.sheet_properties = copy(
        source.sheet_properties
    )

    target.page_margins = copy(
        source.page_margins
    )

    target.page_setup = copy(
        source.page_setup
    )

    target.print_options = copy(
        source.print_options
    )

    target.freeze_panes = (
        source.freeze_panes
    )

    target.sheet_view.showGridLines = (
        source.sheet_view.showGridLines
    )

    for merged_range in (
        source.merged_cells.ranges
    ):
        target.merge_cells(
            str(merged_range)
        )


def ensure_detail_template(
    workbook: Any,
    main_title: str,
    preferred_title: str,
    source_path: Path | None = None,
) -> str:
    """
    建立可供所有公司複製的隱藏明細範本。

    新輸出檔會優先複製來源 Excel 的 155598 工作表，
    因此字型、底色、框線、列高、頁面設定可直接沿用。
    """
    if (
        DETAIL_TEMPLATE_SHEET_NAME
        in workbook.sheetnames
    ):
        template = workbook[
            DETAIL_TEMPLATE_SHEET_NAME
        ]
        template.sheet_state = "veryHidden"
        return template.title

    candidate = find_detail_template_candidate(
        workbook,
        main_title,
        preferred_title,
    )

    if candidate is not None:
        template = workbook.copy_worksheet(
            candidate
        )
        template.title = (
            DETAIL_TEMPLATE_SHEET_NAME
        )
    else:
        imported = False

        if (
            source_path is not None
            and source_path.exists()
        ):
            source_workbook = None

            try:
                source_workbook = (
                    load_workbook(
                        source_path
                    )
                )

                source_main = (
                    main_title
                    if main_title
                    in source_workbook.sheetnames
                    else source_workbook
                    .worksheets[0]
                    .title
                )

                source_candidate = (
                    find_detail_template_candidate(
                        source_workbook,
                        source_main,
                        preferred_title,
                    )
                )

                if source_candidate is not None:
                    template = (
                        workbook.create_sheet(
                            DETAIL_TEMPLATE_SHEET_NAME
                        )
                    )

                    copy_external_detail_template(
                        source_candidate,
                        template,
                    )

                    imported = True

            finally:
                if source_workbook is not None:
                    source_workbook.close()

        if not imported:
            template = workbook.create_sheet(
                DETAIL_TEMPLATE_SHEET_NAME
            )
            build_fallback_detail_template(
                template
            )

    template.sheet_state = "veryHidden"

    return template.title


def copy_template_row_style(
    template: Any,
    template_row: int,
    target: Any,
    target_row: int,
) -> None:
    for column in range(1, 8):
        source_cell = template.cell(
            template_row,
            column,
        )

        target_cell = target.cell(
            target_row,
            column,
        )

        if source_cell.has_style:
            target_cell._style = copy(
                source_cell._style
            )

        target_cell.number_format = (
            source_cell.number_format
        )

        target_cell.alignment = copy(
            source_cell.alignment
        )

        target_cell.protection = copy(
            source_cell.protection
        )

    source_dimension = (
        template.row_dimensions[
            template_row
        ]
    )

    target_dimension = (
        target.row_dimensions[
            target_row
        ]
    )

    target_dimension.height = (
        source_dimension.height
    )

    target_dimension.hidden = (
        source_dimension.hidden
    )

    target_dimension.outlineLevel = (
        source_dimension.outlineLevel
    )


def display_width_units(value: Any) -> int:
    text = normalize_text(value)

    units = 0

    for character in text:
        units += (
            1
            if ord(character) < 128
            else 2
        )

    return units


def estimate_legacy_row_height(
    row: Sequence[Any],
    kind: str,
) -> float:
    """
    155598 使用約 8.43 欄寬與 6 pt 字型。

    每一行約可容納 14 個顯示寬度單位；
    來源檔列高以 8 pt 為一級：
        16.5、24.5、32.5、40.5、...
    """
    if kind == TABLE_KIND_UNPAID:
        values = [
            row[1],
            row[2],
            row[4],
        ]
    else:
        values = [
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
        ]

    line_count = 1

    for value in values:
        units = display_width_units(
            value
        )

        estimated = max(
            1,
            (units + 13) // 14,
        )

        line_count = max(
            line_count,
            estimated,
        )

    return max(
        16.5,
        float(
            line_count * 8
            + 0.5
        ),
    )


def clear_copied_detail_sheet(
    worksheet: Any,
) -> None:
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None
            cell.hyperlink = None
            cell.comment = None

    worksheet.freeze_panes = None

    try:
        worksheet.auto_filter.ref = None
    except Exception:
        pass


def collect_outcome_records(
    outcome: QueryOutcome,
    kind: str,
) -> list[DetailRecord]:
    """
    依實際擷取順序保留每一頁的每一筆資料。

    不再自動刪除相同表格或相同資料；任何重複都必須先在
    validate_page_audits() 被發現並阻止寫入 Excel。
    """
    records: list[DetailRecord] = []

    for table in outcome.tables:
        if table.kind != kind:
            continue

        ensure_table_row_keys(table)

        for index, row in enumerate(table.rows):
            record_key = (
                table.row_keys[index]
                if index < len(table.row_keys)
                else record_fingerprint(row)
            )
            records.append(
                DetailRecord(
                    kind=table.kind,
                    row=list(row),
                    section=(
                        table.section
                        or (
                            QUERY_SECTION_PAID
                            if table.kind == TABLE_KIND_PAID
                            else QUERY_SECTION_CURRENT
                        )
                    ),
                    category=(
                        table.category
                        or table.title
                    ),
                    page_number=table.page_number,
                    total_pages=table.total_pages,
                    declared_total_records=int(
                        table.declared_total_records
                        if table.declared_total_records
                        is not None
                        else table.record_count
                    ),
                    banner_text=(
                        table.banner_text
                        or (
                            f"{table.page_number} / "
                            f"{table.total_pages} 頁，共 "
                            f"{table.declared_total_records or table.record_count} 筆資料"
                        )
                    ),
                    record_key=record_key,
                )
            )

    return records


def collect_outcome_rows(
    outcome: QueryOutcome,
    kind: str,
) -> list[list[Any]]:
    """保留舊呼叫介面；新程式應使用 collect_outcome_records()。"""
    return [
        record.row
        for record in collect_outcome_records(
            outcome,
            kind,
        )
    ]

def write_detail_sheet(
    workbook: Any,
    template_sheet_name: str,
    company: CompanyRow,
    outcome: QueryOutcome,
    queried_at: datetime,
) -> None:
    """
    建立公司違規明細工作表。

    A:G 保留原本 155598 格式；H:N 直接寫入每一筆資料的來源階段、
    類別、擷取頁碼、總頁數、網站宣告總筆數、showbanner 文字與
    唯一鍵，讓使用者可逐筆追查該資料來自哪一頁。
    """
    del queried_at

    remove_existing_detail_variants(
        workbook,
        company.raw_id,
        company.query_id,
    )

    template = workbook[template_sheet_name]
    worksheet = workbook.copy_worksheet(
        template
    )
    worksheet.title = detail_sheet_name(
        company.raw_id,
        company.query_id,
    )
    worksheet.sheet_state = "visible"
    clear_copied_detail_sheet(worksheet)

    unpaid_records = collect_outcome_records(
        outcome,
        TABLE_KIND_UNPAID,
    )
    paid_records = collect_outcome_records(
        outcome,
        TABLE_KIND_PAID,
    )
    all_records = unpaid_records + paid_records

    if not all_records:
        workbook.remove(worksheet)
        return

    current_row = 1
    normal_unpaid_index = 0

    for record_index, record in enumerate(
        all_records
    ):
        row = record.row

        if record.kind == TABLE_KIND_UNPAID:
            status = normalize_text(row[0])
            is_needs_appearance = "需到案" in status

            if is_needs_appearance:
                template_row = 5
            else:
                template_row = (
                    1
                    if normal_unpaid_index == 0
                    else 2
                )
                normal_unpaid_index += 1
        else:
            paid_index = record_index - len(
                unpaid_records
            )
            if paid_index == 0:
                template_row = 6
            elif paid_index == len(paid_records) - 1:
                template_row = 11
            else:
                template_row = 7

        copy_template_row_style(
            template,
            template_row,
            worksheet,
            current_row,
        )

        canonical = list(row[:7])
        canonical += [None] * (
            7 - len(canonical)
        )

        if record.kind == TABLE_KIND_UNPAID:
            canonical[6] = None

        for column, value in enumerate(
            canonical,
            start=1,
        ):
            worksheet.cell(
                current_row,
                column,
                value,
            )

        if record.kind == TABLE_KIND_UNPAID:
            view_cell = worksheet.cell(
                current_row,
                6,
            )
            if not normalize_text(view_cell.value):
                view_cell.value = "檢視"
            if outcome.result_url:
                view_cell.hyperlink = outcome.result_url

        metadata_values = [
            f"查詢階段：{record.section}",
            f"資料分類：{record.category}",
            f"擷取頁碼：{record.page_number}",
            f"總頁數：{record.total_pages}",
            (
                "網站宣告總筆數："
                f"{record.declared_total_records}"
            ),
            f"分頁資訊：{record.banner_text}",
            f"資料唯一鍵：{record.record_key}",
        ]

        source_fill = copy(
            worksheet.cell(current_row, 7).fill
        )

        for offset, value in enumerate(
            metadata_values,
            start=DETAIL_AUDIT_FIRST_COLUMN,
        ):
            cell = worksheet.cell(
                current_row,
                offset,
                value,
            )
            cell.font = Font(
                name="Arial",
                size=7,
                color=LEGACY_PAID_FONT_COLOR,
            )
            cell.fill = copy(source_fill)
            cell.border = legacy_border(
                left=False,
                top=(current_row == 1),
            )
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

        base_height = estimate_legacy_row_height(
            canonical,
            record.kind,
        )
        worksheet.row_dimensions[
            current_row
        ].height = max(base_height, 24.5)
        current_row += 1

    total_rows = len(all_records)

    if worksheet.max_row > total_rows:
        worksheet.delete_rows(
            total_rows + 1,
            worksheet.max_row - total_rows,
        )

    audit_widths = {
        "H": 34,
        "I": 24,
        "J": 15,
        "K": 15,
        "L": 22,
        "M": 34,
        "N": 68,
    }

    for column_letter, width in audit_widths.items():
        worksheet.column_dimensions[
            column_letter
        ].width = width

    worksheet.sheet_view.showGridLines = True

def ensure_log_sheet(workbook: Any) -> Any:
    title = "違規查詢紀錄"
    headers = [
        "查詢時間",
        "Excel列",
        "統一編號",
        "登記名稱",
        "狀態",
        "違規筆數",
        "訊息",
        "結果網址",
        "分頁驗證摘要",
        "重複資料檢查",
    ]

    if title in workbook.sheetnames:
        worksheet = workbook[title]
    else:
        worksheet = workbook.create_sheet(title)

    for column, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(1, column, header)
        cell.fill = TITLE_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(
            left=THIN_GRAY,
            right=THIN_GRAY,
            top=THIN_GRAY,
            bottom=THIN_GRAY,
        )

    worksheet.freeze_panes = "A2"
    widths = [
        20,
        10,
        14,
        32,
        12,
        12,
        58,
        55,
        80,
        24,
    ]

    for index, width in enumerate(
        widths,
        start=1,
    ):
        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = width

    return worksheet

def append_log(
    worksheet: Any,
    company: CompanyRow,
    outcome: QueryOutcome,
    queried_at: datetime,
) -> None:
    worksheet.append(
        [
            queried_at.strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
            company.excel_row,
            company.query_id,
            company.name,
            outcome.status,
            outcome.record_count,
            outcome.message,
            outcome.result_url,
            build_pagination_summary(
                outcome.page_audits
            ),
            outcome_duplicate_text(outcome),
        ]
    )

    worksheet.cell(
        worksheet.max_row,
        3,
    ).number_format = "@"

    for cell in worksheet[worksheet.max_row]:
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )
        cell.border = Border(
            left=THIN_GRAY,
            right=THIN_GRAY,
            top=THIN_GRAY,
            bottom=THIN_GRAY,
        )

def style_audit_header(
    worksheet: Any,
    headers: Sequence[str],
    widths: Sequence[float],
) -> None:
    for column, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(1, column, header)
        cell.fill = TITLE_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(
            left=THIN_GRAY,
            right=THIN_GRAY,
            top=THIN_GRAY,
            bottom=THIN_GRAY,
        )
        worksheet.column_dimensions[
            get_column_letter(column)
        ].width = widths[column - 1]

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}1"
    )


def ensure_pagination_audit_sheet(
    workbook: Any,
) -> Any:
    headers = [
        "查詢時間",
        "Excel列",
        "統一編號",
        "登記名稱",
        "查詢階段",
        "資料分類",
        "擷取頁碼",
        "總頁數",
        "網站宣告總筆數",
        "本頁擷取筆數",
        "本頁唯一筆數",
        "同頁重複筆數",
        "全組擷取筆數",
        "全組唯一筆數",
        "全組重複筆數",
        "分頁資訊",
        "完整性結果",
        "結果網址",
    ]
    widths = [
        20, 10, 14, 30, 34, 20, 12, 12, 20,
        16, 16, 16, 16, 16, 16, 34, 24, 55,
    ]

    if PAGINATION_AUDIT_SHEET_NAME in workbook.sheetnames:
        worksheet = workbook[
            PAGINATION_AUDIT_SHEET_NAME
        ]
    else:
        worksheet = workbook.create_sheet(
            PAGINATION_AUDIT_SHEET_NAME
        )

    style_audit_header(
        worksheet,
        headers,
        widths,
    )
    return worksheet


def ensure_duplicate_audit_sheet(
    workbook: Any,
) -> Any:
    headers = [
        "查詢時間",
        "Excel列",
        "統一編號",
        "登記名稱",
        "查詢階段",
        "資料分類",
        "資料唯一鍵",
        "出現次數",
        "出現頁碼",
        "檢查結果",
    ]
    widths = [
        20, 10, 14, 30, 34, 20, 68, 14, 24, 24,
    ]

    if DUPLICATE_AUDIT_SHEET_NAME in workbook.sheetnames:
        worksheet = workbook[
            DUPLICATE_AUDIT_SHEET_NAME
        ]
    else:
        worksheet = workbook.create_sheet(
            DUPLICATE_AUDIT_SHEET_NAME
        )

    style_audit_header(
        worksheet,
        headers,
        widths,
    )
    return worksheet


def remove_company_rows_from_sheet(
    worksheet: Any,
    company_id_column: int,
    company_id: str,
) -> None:
    for row in range(
        worksheet.max_row,
        1,
        -1,
    ):
        value = normalize_text(
            worksheet.cell(
                row,
                company_id_column,
            ).value
        )

        if value == company_id:
            worksheet.delete_rows(row, 1)


def replace_company_audit_rows(
    pagination_sheet: Any,
    duplicate_sheet: Any,
    company: CompanyRow,
    outcome: QueryOutcome,
    queried_at: datetime,
) -> None:
    """以本次查詢結果覆蓋該公司的舊分頁稽核紀錄。"""
    remove_company_rows_from_sheet(
        pagination_sheet,
        3,
        company.query_id,
    )
    remove_company_rows_from_sheet(
        duplicate_sheet,
        3,
        company.query_id,
    )

    grouped = group_page_audits(
        outcome.page_audits
    )

    for group, items in grouped.items():
        stats = page_audit_group_stats(items)
        expected_pages = set(
            range(1, int(stats["total_pages"]) + 1)
        )
        passed = (
            set(stats["visited_pages"])
            == expected_pages
            and int(stats["extracted_count"])
            == int(stats["declared_total_records"])
            and int(stats["unique_count"])
            == int(stats["declared_total_records"])
            and int(stats["duplicate_count"]) == 0
        )

        for audit in items:
            pagination_sheet.append(
                [
                    queried_at.strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),
                    company.excel_row,
                    company.query_id,
                    company.name,
                    audit.section,
                    audit.category,
                    audit.page_number,
                    audit.total_pages,
                    audit.declared_total_records,
                    audit.extracted_count,
                    audit.unique_count,
                    audit.duplicate_count,
                    stats["extracted_count"],
                    stats["unique_count"],
                    stats["duplicate_count"],
                    audit.banner_text,
                    "通過" if passed else "未通過",
                    audit.result_url,
                ]
            )

            for cell in pagination_sheet[
                pagination_sheet.max_row
            ]:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
                cell.border = Border(
                    left=THIN_GRAY,
                    right=THIN_GRAY,
                    top=THIN_GRAY,
                    bottom=THIN_GRAY,
                )

        page_map: dict[str, list[int]] = defaultdict(list)
        counts: Counter[str] = Counter()

        for audit in items:
            for key in audit.record_keys:
                counts[key] += 1
                page_map[key].append(
                    audit.page_number
                )

        for key, count in counts.items():
            if count <= 1:
                continue

            duplicate_sheet.append(
                [
                    queried_at.strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),
                    company.excel_row,
                    company.query_id,
                    company.name,
                    group[0],
                    group[1],
                    key,
                    count,
                    ", ".join(
                        str(page_number)
                        for page_number in page_map[key]
                    ),
                    "重複擷取，完整性未通過",
                ]
            )

            for cell in duplicate_sheet[
                duplicate_sheet.max_row
            ]:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
                cell.border = Border(
                    left=THIN_GRAY,
                    right=THIN_GRAY,
                    top=THIN_GRAY,
                    bottom=THIN_GRAY,
                )

    pagination_sheet.auto_filter.ref = (
        f"A1:R{max(1, pagination_sheet.max_row)}"
    )
    duplicate_sheet.auto_filter.ref = (
        f"A1:J{max(1, duplicate_sheet.max_row)}"
    )


def atomic_save(
    workbook: Any,
    output_path: Path,
) -> None:
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with tempfile.NamedTemporaryFile(
        prefix=f".{output_path.stem}_",
        suffix=".xlsx",
        dir=output_path.parent,
        delete=False,
    ) as handle:
        temp_path = Path(handle.name)

    try:
        workbook.save(temp_path)

        check = load_workbook(
            temp_path,
            read_only=True,
            data_only=False,
        )
        check.close()

        os.replace(
            temp_path,
            output_path,
        )
    finally:
        if temp_path.exists():
            temp_path.unlink(
                missing_ok=True
            )


def locator_is_usable(locator: Locator) -> bool:
    try:
        return (
            locator.count() > 0
            and locator.first.is_visible()
            and locator.first.is_enabled()
        )
    except PlaywrightError:
        return False


def first_usable(
    locators: Iterable[Locator],
) -> Locator:
    for locator in locators:
        if locator_is_usable(locator):
            return locator.first

    raise RuntimeError(
        "找不到可用的網頁欄位；網站版面可能已改版"
    )


def locate_company_id_input(
    page: Page,
) -> Locator:
    label_patterns = [
        re.compile(r"統一編號", re.I),
        re.compile(r"登記編號", re.I),
        re.compile(r"公司.*編號", re.I),
    ]

    candidates: list[Locator] = [
        page.get_by_label(pattern)
        for pattern in label_patterns
    ]

    candidates += [
        page.locator("input[name*='ban' i]"),
        page.locator("input[id*='ban' i]"),
        page.locator("input[name*='tax' i]"),
        page.locator("input[id*='tax' i]"),
        page.locator("input[name*='company' i]"),
        page.locator("input[id*='company' i]"),
    ]

    try:
        return first_usable(candidates)
    except RuntimeError:
        pass

    handle = page.evaluate_handle(
        """
        () => {
          const visible = el => {
            const s = getComputedStyle(el);
            const r = el.getBoundingClientRect();

            return (
              s.visibility !== 'hidden'
              && s.display !== 'none'
              && r.width > 0
              && r.height > 0
            );
          };

          const inputs = [
            ...document.querySelectorAll('input')
          ].filter(el => {
            const type = (
              el.type || 'text'
            ).toLowerCase();

            return (
              visible(el)
              && !el.disabled
              && [
                'text',
                'tel',
                'number',
                ''
              ].includes(type)
            );
          });

          for (const el of inputs) {
            const nearby = [
              el.getAttribute('aria-label') || '',
              el.placeholder || '',
              el.name || '',
              el.id || '',
              el.closest('label')?.innerText || '',
              el.parentElement?.innerText || '',
              el.parentElement
                ?.previousElementSibling
                ?.innerText || ''
            ].join(' ');

            if (
              /(統一編號|登記編號|公司.{0,4}編號)/
                .test(nearby)
              && !/(驗證碼|captcha|verify)/i
                .test(nearby)
            ) {
              return el;
            }
          }

          return null;
        }
        """
    )

    element = handle.as_element()

    if element is None:
        raise RuntimeError(
            "找不到『統一編號／登記編號』輸入欄位"
        )

    return element


def locate_captcha_input(
    page: Page,
) -> Locator:
    candidates: list[Locator] = [
        page.get_by_label(
            re.compile(r"驗證碼", re.I)
        ),
        page.locator(
            "input[name*='captcha' i]"
        ),
        page.locator(
            "input[id*='captcha' i]"
        ),
        page.locator(
            "input[name*='verify' i]"
        ),
        page.locator(
            "input[id*='verify' i]"
        ),
        page.locator(
            "input[name*='valid' i]"
        ),
        page.locator(
            "input[id*='valid' i]"
        ),
    ]

    try:
        return first_usable(candidates)
    except RuntimeError:
        pass

    handle = page.evaluate_handle(
        """
        () => {
          const visible = el => {
            const s = getComputedStyle(el);
            const r = el.getBoundingClientRect();

            return (
              s.visibility !== 'hidden'
              && s.display !== 'none'
              && r.width > 0
              && r.height > 0
            );
          };

          const inputs = [
            ...document.querySelectorAll('input')
          ].filter(el => {
            const type = (
              el.type || 'text'
            ).toLowerCase();

            return (
              visible(el)
              && !el.disabled
              && [
                'text',
                'tel',
                'number',
                ''
              ].includes(type)
            );
          });

          for (const el of inputs) {
            const nearby = [
              el.getAttribute('aria-label') || '',
              el.placeholder || '',
              el.name || '',
              el.id || '',
              el.closest('label')?.innerText || '',
              el.parentElement?.innerText || '',
              el.parentElement
                ?.previousElementSibling
                ?.innerText || ''
            ].join(' ');

            if (
              /(驗證碼|captcha|verify|validation.?code)/i
                .test(nearby)
            ) {
              return el;
            }
          }

          return null;
        }
        """
    )

    element = handle.as_element()

    if element is None:
        raise RuntimeError(
            "找不到驗證碼輸入欄位"
        )

    return element


def locate_captcha_capture_area(
    page: Page,
) -> Locator:
    """
    找到要截圖保存的 CAPTCHA 圖片。

    第一優先使用 CAPTCHA_IMAGE_XPATH，只截取真正的 <img>；
    這可避免整個 td 內的輸入框、標籤及「換一張」文字干擾 OCR。
    若圖片 XPath 因網站微調失效，才使用語意圖片定位；最後才退回
    CAPTCHA_CAPTURE_XPATH 的 td 區塊。
    """
    exact_image = page.locator(
        f"xpath={CAPTCHA_IMAGE_XPATH}"
    )

    try:
        exact_image.first.wait_for(
            state="visible",
            timeout=10000,
        )

        if exact_image.count() > 0:
            return exact_image.first
    except PlaywrightError:
        pass

    image_candidates: list[Locator] = [
        page.get_by_role(
            "img",
            name=re.compile(
                r"驗證碼|驗證用圖|captcha",
                re.I,
            ),
        ),
        page.locator("img[alt*='驗證碼']"),
        page.locator("img[title*='驗證碼']"),
        page.locator("img[src*='captcha' i]"),
        page.locator("img[id*='captcha' i]"),
        page.locator("img[name*='captcha' i]"),
        page.locator("img[src*='verify' i]"),
        page.locator("img[id*='verify' i]"),
        page.locator("img[src*='valid' i]"),
    ]

    for image_candidate in image_candidates:
        try:
            if (
                image_candidate.count() > 0
                and image_candidate.first.is_visible()
            ):
                return image_candidate.first
        except PlaywrightError:
            continue

    all_images = page.locator("img")

    try:
        image_count = all_images.count()
    except PlaywrightError:
        image_count = 0

    for index in range(image_count):
        image = all_images.nth(index)

        try:
            if not image.is_visible():
                continue

            nearby = image.evaluate(
                """
                el => [
                  el.alt || '',
                  el.title || '',
                  el.name || '',
                  el.id || '',
                  el.className || '',
                  el.getAttribute('src') || '',
                  el.parentElement?.innerText || '',
                  el.parentElement
                    ?.previousElementSibling
                    ?.innerText || '',
                  el.closest('tr')?.innerText || ''
                ].join(' ')
                """
            )
        except PlaywrightError:
            continue

        if re.search(
            r"驗證碼|驗證用圖|captcha|verify|validation.?code",
            normalize_text(nearby),
            re.I,
        ):
            return image

    exact_capture_area = page.locator(
        f"xpath={CAPTCHA_CAPTURE_XPATH}"
    )

    if locator_is_usable(exact_capture_area):
        return exact_capture_area.first

    raise RuntimeError(
        "找不到 CAPTCHA 圖片或擷取區塊；"
        "網站版面可能已改版。"
        f"圖片 XPath：{CAPTCHA_IMAGE_XPATH}；"
        f"區塊 XPath：{CAPTCHA_CAPTURE_XPATH}"
    )

def locate_query_button(
    page: Page,
) -> Locator:
    """
    找到監理服務網的查詢按鈕。

    第一優先使用目前頁面的精確 XPath：

        /html/body/table/tbody/tr[2]/td[1]/div[2]/div/div/form/div/a

    該控制項是 <a> 元素，不一定具有標準 button role、文字或
    input[type=submit] 特徵，因此舊版語意定位可能完全找不到。
    精確 XPath 失效時，才使用語意與 DOM 特徵備援。
    """
    exact_button = page.locator(
        f"xpath={QUERY_BUTTON_XPATH}"
    )

    try:
        exact_button.first.wait_for(
            state="attached",
            timeout=10000,
        )

        if exact_button.count() > 0:
            return exact_button.first

    except PlaywrightError:
        pass

    candidates: list[Locator] = [
        page.get_by_role(
            "button",
            name=re.compile(
                r"^(查詢|送出|確定)$"
            ),
        ),
        page.get_by_role(
            "link",
            name=re.compile(
                r"^(查詢|送出|確定)$"
            ),
        ),
        page.locator(
            "input[type='submit']"
            "[value*='查詢']"
        ),
        page.locator(
            "input[type='button']"
            "[value*='查詢']"
        ),
        page.locator(
            "button:has-text('查詢')"
        ),
        page.locator(
            "a:has-text('查詢')"
        ),
        page.locator(
            "form div a"
        ),
    ]

    try:
        return first_usable(candidates)
    except RuntimeError as original_error:
        pass

    # 最後從包含 CAPTCHA 欄位的 form 內找最可能的 <a>／按鈕。
    try:
        handle = page.evaluate_handle(
            """
            () => {
              const visible = el => {
                const style = getComputedStyle(el);
                const rect = el.getBoundingClientRect();

                return (
                  style.visibility !== 'hidden'
                  && style.display !== 'none'
                  && rect.width > 0
                  && rect.height > 0
                );
              };

              const captchaInputs = [
                ...document.querySelectorAll('input')
              ].filter(el => {
                const nearby = [
                  el.getAttribute('aria-label') || '',
                  el.placeholder || '',
                  el.name || '',
                  el.id || '',
                  el.parentElement?.innerText || ''
                ].join(' ');

                return (
                  visible(el)
                  && /(驗證碼|captcha|verify|validation.?code)/i
                    .test(nearby)
                );
              });

              const form = (
                captchaInputs[0]?.closest('form')
                || document.querySelector('form')
              );

              if (!form) {
                return null;
              }

              const controls = [
                ...form.querySelectorAll(
                  'a, button, input[type="submit"], '
                  + 'input[type="button"], [role="button"], [onclick]'
                )
              ].filter(visible);

              for (const el of controls) {
                const descriptor = [
                  el.innerText || '',
                  el.textContent || '',
                  el.getAttribute('value') || '',
                  el.getAttribute('title') || '',
                  el.getAttribute('aria-label') || '',
                  el.id || '',
                  el.className || '',
                  el.getAttribute('href') || '',
                  el.getAttribute('onclick') || ''
                ].join(' ');

                if (/(查詢|送出|確定|query|search|submit)/i.test(descriptor)) {
                  return el;
                }
              }

              // 此頁的查詢按鈕位在 form 直屬 div 裡的 a。
              const directAnchor = form.querySelector(':scope > div > a');

              if (directAnchor && visible(directAnchor)) {
                return directAnchor;
              }

              return controls.find(el => el.tagName === 'A') || null;
            }
            """
        )

        element = handle.as_element()

        if element is not None:
            return element

    except PlaywrightError:
        pass

    raise RuntimeError(
        "找不到查詢按鈕。"
        f"指定 XPath：{QUERY_BUTTON_XPATH}；"
        "網站版面可能已再次改版"
    )


def click_query_button(
    page: Page,
    timeout_ms: int,
) -> str:
    """
    點擊第二階段查詢按鈕。

    網站的 <a onclick="..."> 會觸發表單送出與導頁。Playwright 的
    locator.click() 可能已經成功點擊並完成導頁，卻因等待導頁結束逾時
    而丟出 TimeoutError。這種情況不可再次點擊舊 locator，否則會把
    已成功的查詢誤判成失敗。
    """
    return click_locator_with_fallback(
        page=page,
        locator=locate_query_button(page),
        exact_xpath=QUERY_BUTTON_XPATH,
        timeout_ms=timeout_ms,
        label="交通違規繳納記錄查詢按鈕",
    )

def body_text(page: Page) -> str:
    try:
        return normalize_text(
            page.locator("body").inner_text(
                timeout=5000
            )
        )
    except PlaywrightError:
        return ""


def contains_any(
    text: str,
    patterns: Sequence[str],
) -> bool:
    return any(
        pattern in text
        for pattern in patterns
    )


def normalize_record_component(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, bool):
        return "1" if value else "0"

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return normalize_text(value)


def record_fingerprint(values: Sequence[Any]) -> str:
    payload = [
        normalize_record_component(value)
        for value in values
    ]

    return hashlib.sha256(
        json.dumps(
            payload,
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()


def ensure_table_row_keys(table: ExtractedTable) -> None:
    if len(table.row_keys) == len(table.rows):
        return

    table.row_keys = [
        record_fingerprint(row)
        for row in table.rows
    ]


def pagination_group_key(
    section: str,
    category: str,
) -> tuple[str, str]:
    return (
        normalize_text(section),
        normalize_text(category),
    )


def group_page_audits(
    audits: Sequence[PageAudit],
) -> dict[tuple[str, str], list[PageAudit]]:
    grouped: dict[
        tuple[str, str],
        list[PageAudit],
    ] = defaultdict(list)

    for audit in audits:
        grouped[
            pagination_group_key(
                audit.section,
                audit.category,
            )
        ].append(audit)

    for items in grouped.values():
        items.sort(
            key=lambda item: item.page_number
        )

    return dict(grouped)


def page_audit_group_stats(
    audits: Sequence[PageAudit],
) -> dict[str, Any]:
    if not audits:
        return {
            "section": "",
            "category": "",
            "total_pages": 0,
            "declared_total_records": 0,
            "visited_pages": [],
            "extracted_count": 0,
            "unique_count": 0,
            "duplicate_count": 0,
            "duplicate_keys": {},
        }

    record_keys = [
        record_key
        for audit in audits
        for record_key in audit.record_keys
    ]

    counts = Counter(record_keys)
    duplicate_keys = {
        key: count
        for key, count in counts.items()
        if count > 1
    }

    return {
        "section": audits[0].section,
        "category": audits[0].category,
        "total_pages": audits[0].total_pages,
        "declared_total_records": (
            audits[0].declared_total_records
        ),
        "visited_pages": [
            audit.page_number
            for audit in audits
        ],
        "extracted_count": sum(
            audit.extracted_count
            for audit in audits
        ),
        "unique_count": len(counts),
        "duplicate_count": sum(
            count - 1
            for count in duplicate_keys.values()
        ),
        "duplicate_keys": duplicate_keys,
    }


def validate_page_audits(
    audits: Sequence[PageAudit],
) -> list[str]:
    """
    驗證每一組分頁：
    1. 所有頁碼均有走訪。
    2. 每頁宣告的總頁數與總筆數一致。
    3. 擷取筆數等於網站宣告總筆數。
    4. 唯一筆數等於網站宣告總筆數。
    5. 不得有跨頁或同頁重複擷取。
    """
    if not audits:
        raise DataIntegrityError(
            "沒有任何分頁稽核資料，無法確認是否完整擷取",
            {"audits": []},
        )

    all_duplicate_keys: list[str] = []

    for group, items in group_page_audits(
        audits
    ).items():
        stats = page_audit_group_stats(items)
        total_pages_values = {
            item.total_pages
            for item in items
        }
        total_records_values = {
            item.declared_total_records
            for item in items
        }

        problems: list[str] = []

        if len(total_pages_values) != 1:
            problems.append(
                "不同頁面宣告的總頁數不一致"
            )

        if len(total_records_values) != 1:
            problems.append(
                "不同頁面宣告的總筆數不一致"
            )

        expected_pages = set(
            range(
                1,
                int(stats["total_pages"]) + 1,
            )
        )
        visited_pages = set(
            int(value)
            for value in stats["visited_pages"]
        )

        if visited_pages != expected_pages:
            problems.append(
                "實際走訪頁碼不完整："
                f"預期={sorted(expected_pages)}；"
                f"實際={sorted(visited_pages)}"
            )

        if (
            int(stats["extracted_count"])
            != int(stats["declared_total_records"])
        ):
            problems.append(
                "擷取筆數與網站宣告總筆數不一致："
                f"擷取={stats['extracted_count']}；"
                f"網站={stats['declared_total_records']}"
            )

        if (
            int(stats["unique_count"])
            != int(stats["declared_total_records"])
        ):
            problems.append(
                "唯一資料筆數與網站宣告總筆數不一致："
                f"唯一={stats['unique_count']}；"
                f"網站={stats['declared_total_records']}"
            )

        if int(stats["duplicate_count"]) > 0:
            problems.append(
                "發現重複擷取："
                f"{stats['duplicate_count']} 筆"
            )
            all_duplicate_keys.extend(
                stats["duplicate_keys"].keys()
            )

        if problems:
            raise DataIntegrityError(
                (
                    f"{group[0]}／{group[1]} 分頁完整性驗證失敗："
                    + "；".join(problems)
                ),
                {
                    "group": {
                        "section": group[0],
                        "category": group[1],
                    },
                    "stats": stats,
                    "page_audits": [
                        asdict(item)
                        for item in items
                    ],
                    "problems": problems,
                },
            )

    return sorted(set(all_duplicate_keys))


def build_pagination_summary(
    audits: Sequence[PageAudit],
) -> str:
    summaries: list[str] = []

    for (_, _), items in group_page_audits(
        audits
    ).items():
        stats = page_audit_group_stats(items)
        summaries.append(
            f"{stats['section']}／{stats['category']}："
            f"{stats['total_pages']} 頁，共 "
            f"{stats['declared_total_records']} 筆資料；"
            f"已擷取 {stats['extracted_count']} 筆；"
            f"唯一 {stats['unique_count']} 筆；"
            f"重複 {stats['duplicate_count']} 筆"
        )

    return "｜".join(summaries)


def outcome_group_summary(
    outcome: QueryOutcome,
    section: str,
    category: str,
) -> str:
    key = pagination_group_key(
        section,
        category,
    )
    grouped = group_page_audits(
        outcome.page_audits
    )
    items = grouped.get(key, [])

    if not items:
        return "未取得分頁資訊"

    stats = page_audit_group_stats(items)

    return (
        f"{stats['total_pages']} 頁，共 "
        f"{stats['declared_total_records']} 筆資料；"
        f"已擷取 {stats['extracted_count']} 筆；"
        f"唯一 {stats['unique_count']} 筆；"
        f"重複 {stats['duplicate_count']} 筆"
    )


def outcome_integrity_text(
    outcome: QueryOutcome,
) -> str:
    if outcome.status != STATUS_SUCCESS:
        return "未通過"

    if not outcome.page_audits:
        return "未取得分頁稽核資料"

    return f"{PAGINATION_VERIFIED_MARKER}；全部頁碼與總筆數一致"


def outcome_duplicate_text(
    outcome: QueryOutcome,
) -> str:
    if outcome.duplicate_keys:
        return f"發現 {len(outcome.duplicate_keys)} 組重複資料"

    if outcome.status == STATUS_SUCCESS:
        return "無重複擷取（0 筆）"

    return "未完成檢查"


def read_pagination_info(
    page: Page,
    banner_xpath: str,
    section: str,
    category: str,
) -> PaginationInfo | None:
    """
    讀取目前分頁資訊。

    優先使用指定 XPath；若網站微調 DOM，則尋找 span#showbanner 或
    內含 txtPage/goPage 的 span。只有找到候選節點但內容無法解析時才
    報完整性錯誤；完全沒有候選時回傳 None，交由單頁判斷處理。
    """
    if not page_is_open(page):
        raise RuntimeError(
            "Target page, context or browser has been closed"
        )

    candidates = (
        page.locator(f"xpath={banner_xpath}"),
        page.locator("span#showbanner"),
        page.locator(
            "span:has(input[name='txtPage']), "
            "span:has(input#goPage)"
        ),
    )

    candidate_payloads: list[dict[str, Any]] = []
    seen_handles: set[str] = set()

    for locator in candidates:
        try:
            count = locator.count()
        except PlaywrightError:
            continue

        for index in range(count):
            item = locator.nth(index)

            try:
                if not item.is_visible():
                    continue

                payload = item.evaluate(
                    """
                    element => {
                      const norm = value => (value || '')
                        .replace(/\\s+/g, ' ')
                        .trim();

                      const pageInput = element.querySelector(
                        "input[name='txtPage'], input#goPage"
                      );
                      const totalInput = element.querySelector(
                        "input[name='total'], input#total"
                      );

                      return {
                        key: [
                          element.tagName || '',
                          element.id || '',
                          element.className || '',
                          norm(element.textContent)
                        ].join('|'),
                        text: norm(element.innerText),
                        textContent: norm(element.textContent),
                        currentPage: pageInput?.value || '',
                        totalPages: totalInput?.value || ''
                      };
                    }
                    """
                )
            except PlaywrightError:
                continue

            key = normalize_text(payload.get("key"))
            if key in seen_handles:
                continue
            seen_handles.add(key)
            candidate_payloads.append(payload)

    if not candidate_payloads:
        return None

    parse_errors: list[dict[str, Any]] = []

    for payload in candidate_payloads:
        raw_text = normalize_text(
            payload.get("text")
            or payload.get("textContent")
        )
        current_text = normalize_text(
            payload.get("currentPage")
        )
        total_pages_text = normalize_text(
            payload.get("totalPages")
        )

        total_pages_match = re.search(
            r"/\s*([\d,，]+)\s*頁",
            raw_text,
        )
        total_records_match = re.search(
            r"共\s*([\d,，]+)\s*筆(?:資料)?",
            raw_text,
        )

        if not current_text or not re.fullmatch(
            r"\d+",
            current_text,
        ):
            current_match = re.search(
                r"([\d,，]+)\s*/\s*[\d,，]+\s*頁",
                raw_text,
            )
            current_text = (
                current_match.group(1)
                if current_match
                else ""
            )

        if not total_pages_text or not re.fullmatch(
            r"\d+",
            total_pages_text,
        ):
            total_pages_text = (
                total_pages_match.group(1)
                if total_pages_match
                else ""
            )

        if not (
            current_text
            and total_pages_text
            and total_records_match
        ):
            parse_errors.append(
                {
                    "raw_text": raw_text,
                    "payload": payload,
                }
            )
            continue

        current_page = int(
            re.sub(r"[,，]", "", current_text)
        )
        total_pages = max(
            1,
            int(
                re.sub(
                    r"[,，]",
                    "",
                    total_pages_text,
                )
            ),
        )
        declared_total_records = int(
            re.sub(
                r"[,，]",
                "",
                total_records_match.group(1),
            )
        )

        if not (1 <= current_page <= total_pages):
            parse_errors.append(
                {
                    "current_page": current_page,
                    "total_pages": total_pages,
                    "raw_text": raw_text,
                }
            )
            continue

        banner_text = (
            f"{current_page} / {total_pages} 頁，共 "
            f"{declared_total_records} 筆資料"
        )

        return PaginationInfo(
            section=section,
            category=category,
            current_page=current_page,
            total_pages=total_pages,
            declared_total_records=(
                declared_total_records
            ),
            banner_text=banner_text,
            raw_text=raw_text,
        )

    raise DataIntegrityError(
        f"分頁資訊格式無法解析：{section}／{category}",
        {
            "xpath": banner_xpath,
            "candidates": parse_errors,
            "result_url": safe_page_url(page),
        },
    )

def inspect_pagination_evidence(
    page: Page,
    banner_xpath: str,
) -> dict[str, Any]:
    """
    檢查頁面是否真的存在多頁控制項。

    單頁結果常只有空的 span#pagebanner，沒有 showbanner、goPage、total、
    Go 或上一頁／下一頁。空 pagebanner 不算分頁證據。
    """
    if not page_is_open(page):
        raise RuntimeError(
            "Target page, context or browser has been closed"
        )

    try:
        return page.evaluate(
            """
            xpath => {
              const norm = value => (value || '')
                .replace(/\\s+/g, ' ')
                .trim();

              const exact = document.evaluate(
                xpath,
                document,
                null,
                XPathResult.FIRST_ORDERED_NODE_TYPE,
                null
              ).singleNodeValue;

              const selectors = [
                "input[name='txtPage']",
                "input#goPage",
                "input[name='total']",
                "input#total",
                "a#goUrl",
                "button#goUrl",
                "input[value='Go']",
                "button"
              ];

              const controls = [];
              for (const selector of selectors) {
                for (const element of document.querySelectorAll(selector)) {
                  const text = norm([
                    element.tagName || '',
                    element.id || '',
                    element.getAttribute('name') || '',
                    element.getAttribute('value') || '',
                    element.getAttribute('href') || '',
                    element.innerText || ''
                  ].join(' '));

                  if (
                    selector === 'button'
                    && !/^(Go|下一頁|下頁|上一頁|上頁)$/i.test(
                      norm(element.innerText)
                    )
                  ) {
                    continue;
                  }

                  controls.push(text);
                }
              }

              const bannerTexts = [
                ...document.querySelectorAll(
                  "span#showbanner, span#pagebanner"
                )
              ]
                .map(element => norm(element.textContent))
                .filter(Boolean);

              const paginationText = bannerTexts.find(text =>
                /\\d+\\s*\\/\\s*\\d+\\s*頁/.test(text)
                || /共\\s*\\d+\\s*筆/.test(text)
              ) || '';

              return {
                exactBannerFound: Boolean(exact),
                exactBannerText: norm(exact?.textContent || ''),
                controls,
                bannerTexts,
                paginationText,
                hasEvidence: Boolean(
                  controls.length > 0 || paginationText
                )
              };
            }
            """,
            banner_xpath,
        )
    except PlaywrightError as exc:
        if is_browser_session_closed_exception(exc):
            raise
        return {
            "exactBannerFound": False,
            "exactBannerText": "",
            "controls": [],
            "bannerTexts": [],
            "paginationText": "",
            "hasEvidence": False,
            "inspection_error": (
                f"{type(exc).__name__}: {exc}"
            ),
        }


def build_single_page_info(
    section: str,
    category: str,
    extracted_count: int,
) -> PaginationInfo:
    return PaginationInfo(
        section=section,
        category=category,
        current_page=1,
        total_pages=1,
        declared_total_records=extracted_count,
        banner_text=(
            f"1 / 1 頁，共 {extracted_count} 筆資料"
            "（單頁，網站未顯示分頁列）"
        ),
        raw_text=(
            "單頁結果：網站未產生 showbanner/goPage/total；"
            f"依目標結果表格確認 {extracted_count} 筆"
        ),
    )


def read_scoped_pagination_controls(
    page: Page,
    banner_xpath: str,
) -> dict[str, Any] | None:
    """
    讀取「目前結果表格所屬 pagebar」的分頁控制項。

    網站同一頁可能存在重複的 id，例如：
    - previous
    - next
    - goUrl
    - goPage
    - total

    因此不可使用 document.querySelector() 直接拿整頁第一個控制項。
    本函式先依指定 XPath 找到目前類別的 showbanner，再限制在其所屬
    .pagebar 內讀取上一頁、下一頁與 Go 連結。
    """
    if not page_is_open(page):
        raise RuntimeError(
            "Target page, context or browser has been closed"
        )

    try:
        payload = page.evaluate(
            r"""
            xpath => {
              const norm = value => (value || '')
                .replace(/\s+/g, ' ')
                .trim();

              const visible = element => {
                if (!element) {
                  return false;
                }

                const style = getComputedStyle(element);
                const rect = element.getBoundingClientRect();

                return (
                  style.visibility !== 'hidden'
                  && style.display !== 'none'
                  && rect.width > 0
                  && rect.height > 0
                );
              };

              let exact = document.evaluate(
                xpath,
                document,
                null,
                XPathResult.FIRST_ORDERED_NODE_TYPE,
                null
              ).singleNodeValue;

              if (exact && !visible(exact)) {
                exact = null;
              }

              const semantic = [
                ...document.querySelectorAll(
                  "span#showbanner, span"
                )
              ].find(element => (
                visible(element)
                && element.querySelector(
                  "input[name='txtPage'], input#goPage"
                )
              ));

              const banner = exact || semantic || null;

              if (!banner) {
                return null;
              }

              const pagebar = (
                banner.closest('.pagebar')
                || banner.parentElement
                || null
              );

              if (!pagebar) {
                return null;
              }

              const previous = pagebar.querySelector(
                "a#previous, a[rel='prev']"
              );

              const next = pagebar.querySelector(
                "a#next, a[rel='next']"
              );

              const go = pagebar.querySelector(
                "a#goUrl, button#goUrl, "
                + "a[href*='method=pagination'], "
                + "a[href*='method=nopayPagination']"
              );

              const pageInput = (
                banner.querySelector(
                  "input[name='txtPage'], input#goPage"
                )
                || pagebar.querySelector(
                  "input[name='txtPage'], input#goPage"
                )
              );

              const totalInput = (
                banner.querySelector(
                  "input[name='total'], input#total"
                )
                || pagebar.querySelector(
                  "input[name='total'], input#total"
                )
              );

              const rawHref = element => (
                element?.getAttribute('href') || ''
              );

              const absoluteHref = element => (
                element?.href || ''
              );

              return {
                currentPage: pageInput?.value || '',
                totalPages: totalInput?.value || '',
                previousHref: rawHref(previous),
                previousUrl: absoluteHref(previous),
                nextHref: rawHref(next),
                nextUrl: absoluteHref(next),
                goHref: rawHref(go),
                goUrl: absoluteHref(go),
                bannerText: norm(banner.textContent),
                pagebarText: norm(pagebar.innerText),
                pagebarHtml: pagebar.outerHTML || '',
                location: window.location.href
              };
            }
            """,
            banner_xpath,
        )
    except PlaywrightError as exc:
        if is_browser_session_closed_exception(exc):
            raise

        raise DataIntegrityError(
            "無法讀取目前結果表格的分頁控制項",
            {
                "xpath": banner_xpath,
                "exception": (
                    f"{type(exc).__name__}: {exc}"
                ),
                "result_url": safe_page_url(page),
            },
        ) from exc

    if not isinstance(payload, dict):
        return None

    return {
        str(key): value
        for key, value in payload.items()
    }


def pagination_parameter_from_url(
    value: str,
) -> str:
    """
    從 DisplayTag 分頁網址找出頁碼 query parameter。

    監理服務網目前常見形式：
        d-49440-p
        d-2637073-p

    不把 method、tab 等其他數字欄位誤認成頁碼。
    """
    text = normalize_text(value)

    if not text:
        return ""

    try:
        query_items = parse_qsl(
            urlsplit(text).query,
            keep_blank_values=True,
        )
    except ValueError:
        query_items = []

    for key, _ in query_items:
        if re.search(
            r"(?:^|[-_])p$",
            key,
            re.I,
        ):
            return key

    match = re.search(
        r"[?&]([^=&?#]+(?:-|_)p)="
        r"\d+",
        text,
        re.I,
    )

    return (
        normalize_text(match.group(1))
        if match
        else ""
    )


def pagination_page_from_url(
    value: str,
    parameter: str,
) -> int | None:
    if not value or not parameter:
        return None

    try:
        query_items = parse_qsl(
            urlsplit(value).query,
            keep_blank_values=True,
        )
    except ValueError:
        return None

    for key, raw_value in query_items:
        if key != parameter:
            continue

        compact = re.sub(
            r"[,，\s]",
            "",
            normalize_text(raw_value),
        )

        if compact.isdigit():
            return int(compact)

    return None


def replace_pagination_page_in_url(
    value: str,
    parameter: str,
    target_page: int,
) -> str:
    """
    保留原本 method、其他 query parameter 與 #anchor，
    只替換目前 pagebar 的 DisplayTag 頁碼。
    """
    if not value or not parameter:
        return ""

    parts = urlsplit(value)
    query_items = parse_qsl(
        parts.query,
        keep_blank_values=True,
    )

    replaced = False
    output_items: list[
        tuple[str, str]
    ] = []

    for key, raw_value in query_items:
        if key == parameter:
            output_items.append(
                (
                    key,
                    str(target_page),
                )
            )
            replaced = True
        else:
            output_items.append(
                (
                    key,
                    raw_value,
                )
            )

    if not replaced:
        output_items.append(
            (
                parameter,
                str(target_page),
            )
        )

    return urlunsplit(
        (
            parts.scheme,
            parts.netloc,
            parts.path,
            urlencode(
                output_items,
                doseq=True,
            ),
            parts.fragment,
        )
    )


def build_pagination_target_url(
    current_url: str,
    controls: dict[str, Any],
    target_page: int,
) -> tuple[str, str, str]:
    """
    由目前 pagebar 的上一頁／下一頁連結建立目標網址。

    優先規則：
    1. 下一頁或上一頁連結本身已指向 target_page，直接使用。
    2. 以目前 pagebar 的相鄰頁連結為基底，只替換其頁碼參數。
    3. 最後才使用 Go 連結。

    這裡不依賴 changePage() 產生的 href。實際錯誤紀錄顯示，在第 2 頁
    要前往第 3 頁時，changePage() 產生的 navigation_url 反而是第 1
    頁；但同一個 pagebar 的 next href 正確指向第 3 頁。
    """
    try:
        current_page = int(
            normalize_text(
                controls.get("currentPage")
            )
            or "0"
        )
    except ValueError:
        current_page = 0

    href_entries = [
        (
            "next",
            normalize_text(
                controls.get("nextUrl")
                or controls.get("nextHref")
            ),
        ),
        (
            "previous",
            normalize_text(
                controls.get("previousUrl")
                or controls.get("previousHref")
            ),
        ),
        (
            "go",
            normalize_text(
                controls.get("goUrl")
                or controls.get("goHref")
            ),
        ),
    ]

    absolute_entries: list[
        tuple[str, str]
    ] = []

    for label, href in href_entries:
        if not href or href == "#":
            continue

        absolute_entries.append(
            (
                label,
                urljoin(
                    current_url,
                    href,
                ),
            )
        )

    if target_page == current_page + 1:
        preferred_labels = (
            "next",
            "previous",
            "go",
        )
    elif target_page == current_page - 1:
        preferred_labels = (
            "previous",
            "next",
            "go",
        )
    else:
        preferred_labels = (
            "next",
            "previous",
            "go",
        )

    ordered_entries = sorted(
        absolute_entries,
        key=lambda item: (
            preferred_labels.index(
                item[0]
            )
            if item[0] in preferred_labels
            else len(preferred_labels)
        ),
    )

    # 相鄰頁連結若已經正確指向目標頁，完全不修改它。
    for label, absolute_url in ordered_entries:
        parameter = pagination_parameter_from_url(
            absolute_url
        )

        if not parameter:
            continue

        page_number = pagination_page_from_url(
            absolute_url,
            parameter,
        )

        if page_number == target_page:
            return (
                absolute_url,
                parameter,
                f"{label} 連結直接指向目標頁",
            )

    # 找到目前 pagebar 真正使用的頁碼參數。
    parameter = ""

    for _, absolute_url in ordered_entries:
        parameter = pagination_parameter_from_url(
            absolute_url
        )

        if parameter:
            break

    if not parameter:
        parameter = pagination_parameter_from_url(
            current_url
        )

    if not parameter:
        return "", "", "找不到 DisplayTag 頁碼參數"

    for label, absolute_url in ordered_entries:
        target_url = replace_pagination_page_in_url(
            absolute_url,
            parameter,
            target_page,
        )

        if target_url:
            return (
                target_url,
                parameter,
                f"以 {label} 連結為基底替換頁碼",
            )

    return "", parameter, "pagebar 沒有可用連結"


def pagination_page_signature(
    page: Page,
) -> str:
    """
    取得目前頁面實際顯示內容的簽章。

    此簽章只用來判斷頁面是否已穩定，不再把「內容必須與上一頁不同」
    當成導頁成功的唯一依據。是否真正到達目標頁，改由：
    - 伺服器回傳後的 URL query parameter；
    - showbanner/goPage 的目前頁碼
    共同確認。

    若網站真的在兩頁回傳重複紀錄，後續 record key 稽核仍會攔截。
    """
    text = body_text(page)

    if not text:
        return ""

    return hashlib.sha256(
        text.encode("utf-8")
    ).hexdigest()


def wait_for_pagination_page(
    page: Page,
    banner_xpath: str,
    section: str,
    category: str,
    target_page: int,
    before_signature: str,
    timeout_ms: int,
    expected_parameter: str = "",
) -> bool:
    """
    等待伺服器真正回傳目標頁。

    驗證條件：
    1. showbanner/goPage 顯示 target_page。
    2. 若已知 DisplayTag 參數，最終 URL 中該參數也必須是 target_page。
    3. 目前 DOM 連續兩次保持穩定。

    不再只看 input.value，也不強迫資料一定與上一頁不同；重複資料由
    validate_page_audits() 負責檢出。
    """
    del before_signature

    deadline = time.monotonic() + min(
        max(timeout_ms, 1000),
        30000,
    ) / 1000.0

    stable_count = 0
    last_signature = ""

    while time.monotonic() < deadline:
        if not page_is_open(page):
            raise RuntimeError(
                "Target page, context or browser has been closed"
            )

        try:
            page.wait_for_load_state(
                "domcontentloaded",
                timeout=500,
            )
        except PlaywrightTimeoutError:
            pass
        except PlaywrightError as exc:
            if is_browser_session_closed_exception(exc):
                raise

        try:
            info = read_pagination_info(
                page,
                banner_xpath,
                section,
                category,
            )
        except DataIntegrityError:
            info = None

        url_page = (
            pagination_page_from_url(
                safe_page_url(page),
                expected_parameter,
            )
            if expected_parameter
            else None
        )

        url_matches = (
            url_page == target_page
            if expected_parameter
            else True
        )

        if (
            info is not None
            and info.current_page == target_page
            and url_matches
        ):
            signature = pagination_page_signature(
                page
            )

            if signature:
                if signature == last_signature:
                    stable_count += 1
                else:
                    last_signature = signature
                    stable_count = 1

                if stable_count >= 2:
                    return True
        else:
            stable_count = 0
            last_signature = ""

        time.sleep(0.25)

    return False


def navigate_to_pagination_page(
    page: Page,
    banner_xpath: str,
    section: str,
    category: str,
    target_page: int,
    timeout_ms: int,
) -> None:
    """
    使用目前 pagebar 的「正確相鄰頁連結」切換分頁。

    本版不再信任 changePage(input) 產生的 goUrl。實際執行紀錄顯示：
    - 當前頁：2
    - 目標頁：3
    - next href：第 3 頁
    - changePage 產生的 navigation_url：第 1 頁

    因此現在直接使用該 pagebar 的 next／previous href，或以其網址中的
    DisplayTag 參數建立 target_page URL，再由 page.goto() 向伺服器
    載入。這可避免第 2 頁與第 1 頁之間來回跳轉。
    """
    current_info = read_pagination_info(
        page,
        banner_xpath,
        section,
        category,
    )

    if current_info is None:
        raise DataIntegrityError(
            f"找不到分頁資訊，無法前往第 {target_page} 頁："
            f"{section}／{category}",
            {
                "xpath": banner_xpath,
                "result_url": safe_page_url(page),
            },
        )

    if current_info.current_page == target_page:
        return

    if not (
        1
        <= target_page
        <= current_info.total_pages
    ):
        raise DataIntegrityError(
            f"要求前往的頁碼超出範圍：第 {target_page} 頁",
            {
                "section": section,
                "category": category,
                "current_page": (
                    current_info.current_page
                ),
                "total_pages": (
                    current_info.total_pages
                ),
            },
        )

    before_url = safe_page_url(page)
    before_signature = pagination_page_signature(
        page
    )
    errors: list[str] = []

    controls = read_scoped_pagination_controls(
        page,
        banner_xpath,
    )

    if controls is None:
        raise DataIntegrityError(
            "已讀到分頁資訊，但找不到其所屬 pagebar 控制項",
            {
                "xpath": banner_xpath,
                "target_page": target_page,
                "result_url": before_url,
            },
        )

    (
        target_url,
        page_parameter,
        target_url_source,
    ) = build_pagination_target_url(
        before_url,
        controls,
        target_page,
    )

    if target_url:
        try:
            response = page.goto(
                target_url,
                wait_until="domcontentloaded",
                timeout=timeout_ms,
            )

            if response is not None:
                status = int(response.status)

                if status >= 400:
                    errors.append(
                        "目標分頁 HTTP 狀態："
                        f"{status}"
                    )

        except PlaywrightTimeoutError as exc:
            errors.append(
                "前往目標分頁 URL："
                f"{type(exc).__name__}: {exc}"
            )
        except PlaywrightError as exc:
            if is_browser_session_closed_exception(exc):
                raise

            errors.append(
                "前往目標分頁 URL："
                f"{type(exc).__name__}: {exc}"
            )

        if wait_for_pagination_page(
            page=page,
            banner_xpath=banner_xpath,
            section=section,
            category=category,
            target_page=target_page,
            before_signature=before_signature,
            timeout_ms=min(
                timeout_ms,
                15000,
            ),
            expected_parameter=page_parameter,
        ):
            return

    # page.goto() 未通過時，才點擊同一個 pagebar 內的相鄰頁連結。
    try:
        banner_locator = page.locator(
            f"xpath={banner_xpath}"
        ).first

        if banner_locator.count() <= 0:
            banner_locator = page.locator(
                "span#showbanner:visible"
            ).first

        pagebar = banner_locator.locator(
            "xpath=ancestor::*"
            "[contains(concat(' ', normalize-space(@class), ' '), "
            "' pagebar ')][1]"
        )

        current_page_now = read_pagination_info(
            page,
            banner_xpath,
            section,
            category,
        )

        current_number = (
            current_page_now.current_page
            if current_page_now is not None
            else current_info.current_page
        )

        if target_page > current_number:
            adjacent = pagebar.locator(
                "a#next, a[rel='next']"
            ).first
            adjacent_label = "下一頁"
        else:
            adjacent = pagebar.locator(
                "a#previous, a[rel='prev']"
            ).first
            adjacent_label = "上一頁"

        if adjacent.count() > 0:
            adjacent_href = normalize_text(
                adjacent.get_attribute(
                    "href"
                )
            )
            adjacent_url = (
                urljoin(
                    safe_page_url(page),
                    adjacent_href,
                )
                if adjacent_href
                else ""
            )
            adjacent_parameter = (
                pagination_parameter_from_url(
                    adjacent_url
                )
            )

            # 相鄰連結未必直接等於目標頁；必要時只替換其頁碼。
            if adjacent_url and adjacent_parameter:
                adjacent_url = (
                    replace_pagination_page_in_url(
                        adjacent_url,
                        adjacent_parameter,
                        target_page,
                    )
                )

            if adjacent_url:
                try:
                    page.goto(
                        adjacent_url,
                        wait_until="domcontentloaded",
                        timeout=timeout_ms,
                    )
                except PlaywrightTimeoutError as exc:
                    errors.append(
                        f"{adjacent_label}連結導頁："
                        f"{type(exc).__name__}: {exc}"
                    )
                except PlaywrightError as exc:
                    if is_browser_session_closed_exception(exc):
                        raise

                    errors.append(
                        f"{adjacent_label}連結導頁："
                        f"{type(exc).__name__}: {exc}"
                    )

                if wait_for_pagination_page(
                    page=page,
                    banner_xpath=banner_xpath,
                    section=section,
                    category=category,
                    target_page=target_page,
                    before_signature=before_signature,
                    timeout_ms=min(
                        timeout_ms,
                        15000,
                    ),
                    expected_parameter=(
                        adjacent_parameter
                    ),
                ):
                    return

    except PlaywrightError as exc:
        if is_browser_session_closed_exception(exc):
            raise

        errors.append(
            "相鄰頁連結備援："
            f"{type(exc).__name__}: {exc}"
        )

    final_info = read_pagination_info(
        page,
        banner_xpath,
        section,
        category,
    )

    raise DataIntegrityError(
        f"無法由伺服器載入第 {target_page} 頁："
        f"{section}／{category}",
        {
            "xpath": banner_xpath,
            "target_page": target_page,
            "current_page": (
                final_info.current_page
                if final_info is not None
                else None
            ),
            "total_pages": (
                current_info.total_pages
            ),
            "before_url": before_url,
            "target_url": target_url,
            "target_url_source": (
                target_url_source
            ),
            "page_parameter": (
                page_parameter
            ),
            "result_url": safe_page_url(page),
            "controls": controls,
            "errors": errors,
        },
    )


def attach_pagination_to_tables(
    tables: Sequence[ExtractedTable],
    info: PaginationInfo,
) -> list[ExtractedTable]:
    output: list[ExtractedTable] = []

    for table in tables:
        table.section = info.section
        table.category = info.category
        table.page_number = info.current_page
        table.total_pages = info.total_pages
        table.declared_total_records = (
            info.declared_total_records
        )
        table.banner_text = info.banner_text
        ensure_table_row_keys(table)
        output.append(table)

    return output


def capture_paginated_group(
    page: Page,
    *,
    banner_xpath: str,
    section: str,
    category: str,
    extractor: Any,
    timeout_ms: int,
    max_pages: int = 1000,
) -> tuple[list[ExtractedTable], list[PageAudit]]:
    """
    逐頁擷取單一結果類別，並驗證頁碼、總筆數與重複資料。

    單頁結果的網站不一定輸出 showbanner/goPage/total。這時只有在：
    - 已成功擷取目標資料表；且
    - 頁面完全沒有任何分頁控制證據
    才會建立 1/1 的稽核資料。若仍存在 goPage、total 或 Go，則絕不
    擅自判定單頁，以免再次漏抓。
    """
    initial_info = read_pagination_info(
        page,
        banner_xpath,
        section,
        category,
    )

    if initial_info is None:
        tables = list(extractor())
        extracted_count = sum(
            table.record_count
            for table in tables
        )
        text = body_text(page)
        evidence = inspect_pagination_evidence(
            page,
            banner_xpath,
        )

        if (
            extracted_count == 0
            and contains_any(text, NO_DATA_PATTERNS)
        ):
            synthetic = build_single_page_info(
                section,
                category,
                0,
            )
            audit = PageAudit(
                section=section,
                category=category,
                page_number=1,
                total_pages=1,
                declared_total_records=0,
                banner_text=synthetic.banner_text,
                extracted_count=0,
                record_keys=[],
                result_url=safe_page_url(page),
            )
            validate_page_audits([audit])
            return [], [audit]

        if extracted_count > 0 and not bool(
            evidence.get("hasEvidence")
        ):
            synthetic = build_single_page_info(
                section,
                category,
                extracted_count,
            )
            tables = attach_pagination_to_tables(
                tables,
                synthetic,
            )
            record_keys = [
                record_key
                for table in tables
                for record_key in table.row_keys
            ]
            audit = PageAudit(
                section=section,
                category=category,
                page_number=1,
                total_pages=1,
                declared_total_records=extracted_count,
                banner_text=synthetic.banner_text,
                extracted_count=extracted_count,
                record_keys=record_keys,
                result_url=safe_page_url(page),
            )
            validate_page_audits([audit])
            return tables, [audit]

        raise DataIntegrityError(
            f"找不到可解析的分頁資訊，無法驗證完整筆數："
            f"{section}／{category}",
            {
                "xpath": banner_xpath,
                "extracted_count": extracted_count,
                "pagination_evidence": evidence,
                "body_excerpt": text[:2000],
                "result_url": safe_page_url(page),
            },
        )

    if initial_info.total_pages > max_pages:
        raise DataIntegrityError(
            f"網站宣告頁數 {initial_info.total_pages} 超過安全上限 {max_pages}",
            asdict(initial_info),
        )

    if initial_info.current_page != 1:
        navigate_to_pagination_page(
            page=page,
            banner_xpath=banner_xpath,
            section=section,
            category=category,
            target_page=1,
            timeout_ms=timeout_ms,
        )

    all_tables: list[ExtractedTable] = []
    audits: list[PageAudit] = []

    for target_page in range(
        1,
        initial_info.total_pages + 1,
    ):
        if target_page > 1:
            navigate_to_pagination_page(
                page=page,
                banner_xpath=banner_xpath,
                section=section,
                category=category,
                target_page=target_page,
                timeout_ms=timeout_ms,
            )

        info = read_pagination_info(
            page,
            banner_xpath,
            section,
            category,
        )

        if info is None:
            raise DataIntegrityError(
                f"第 {target_page} 頁找不到分頁資訊："
                f"{section}／{category}",
                {"xpath": banner_xpath},
            )

        if info.current_page != target_page:
            raise DataIntegrityError(
                f"頁碼切換失敗：預期第 {target_page} 頁，"
                f"實際第 {info.current_page} 頁",
                asdict(info),
            )

        if (
            info.total_pages != initial_info.total_pages
            or info.declared_total_records
            != initial_info.declared_total_records
        ):
            raise DataIntegrityError(
                "不同頁面顯示的總頁數或總筆數不一致",
                {
                    "initial": asdict(initial_info),
                    "current": asdict(info),
                },
            )

        extraction_deadline = time.monotonic() + min(
            max(timeout_ms, 1000),
            12000,
        ) / 1000.0
        page_tables: list[ExtractedTable] = []
        page_keys: list[str] = []
        extracted_count = 0
        previous_page_keys = (
            audits[-1].record_keys
            if audits
            else []
        )

        while time.monotonic() < extraction_deadline:
            page_tables = attach_pagination_to_tables(
                list(extractor()),
                info,
            )
            page_keys = [
                record_key
                for table in page_tables
                for record_key in table.row_keys
            ]
            extracted_count = sum(
                table.record_count
                for table in page_tables
            )

            nonempty_ready = (
                info.declared_total_records == 0
                or extracted_count > 0
            )
            not_stale_previous_page = not (
                target_page > 1
                and page_keys
                and page_keys == previous_page_keys
            )

            if nonempty_ready and not_stale_previous_page:
                break

            time.sleep(0.25)

        if (
            info.declared_total_records > 0
            and extracted_count <= 0
        ):
            raise DataIntegrityError(
                f"網站宣告共 {info.declared_total_records} 筆，"
                f"但第 {target_page} 頁沒有擷取到任何資料",
                {
                    "pagination": asdict(info),
                    "result_url": safe_page_url(page),
                },
            )

        if (
            target_page > 1
            and page_keys
            and page_keys == previous_page_keys
        ):
            raise DataIntegrityError(
                f"第 {target_page} 頁內容仍與前一頁完全相同，"
                "可能尚未切頁或發生重複擷取",
                {
                    "pagination": asdict(info),
                    "previous_page_keys": previous_page_keys,
                    "current_page_keys": page_keys,
                    "result_url": safe_page_url(page),
                },
            )

        audits.append(
            PageAudit(
                section=section,
                category=category,
                page_number=target_page,
                total_pages=info.total_pages,
                declared_total_records=(
                    info.declared_total_records
                ),
                banner_text=info.banner_text,
                extracted_count=extracted_count,
                record_keys=page_keys,
                result_url=safe_page_url(page),
            )
        )
        all_tables.extend(page_tables)

    validate_page_audits(audits)

    return all_tables, audits

def json_safe(value: Any) -> Any:
    if value is None or isinstance(
        value,
        (str, int, float, bool),
    ):
        return value

    if isinstance(value, datetime):
        return value.isoformat()

    if isinstance(value, Path):
        return str(value)

    if hasattr(value, "__dataclass_fields__"):
        return {
            key: json_safe(item)
            for key, item in asdict(value).items()
        }

    if isinstance(value, dict):
        return {
            str(key): json_safe(item)
            for key, item in value.items()
        }

    if isinstance(value, (list, tuple, set)):
        return [
            json_safe(item)
            for item in value
        ]

    return normalize_text(value)


def write_error_json(
    error_log_dir: Path,
    company: CompanyRow,
    phase: str,
    message: str,
    *,
    page: Page | None = None,
    exception: BaseException | None = None,
    outcome: QueryOutcome | None = None,
    details: dict[str, Any] | None = None,
) -> Path:
    error_log_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    timestamp = datetime.now()
    safe_phase = re.sub(
        r"[^A-Za-z0-9_-]",
        "_",
        phase,
    )[:60]
    filename = (
        f"{company.query_id}_"
        f"{timestamp.strftime('%Y%m%d_%H%M%S_%f')}_"
        f"{safe_phase}.json"
    )
    path = error_log_dir / filename

    payload: dict[str, Any] = {
        "timestamp": timestamp.isoformat(),
        "phase": phase,
        "company": asdict(company),
        "message": message,
        "result_url": (
            safe_page_url(page)
            if page is not None
            else (
                outcome.result_url
                if outcome is not None
                else ""
            )
        ),
        "details": details or {},
    }

    if exception is not None:
        payload["exception"] = {
            "type": type(exception).__name__,
            "message": str(exception),
            "traceback": "".join(
                traceback.format_exception(
                    type(exception),
                    exception,
                    exception.__traceback__,
                )
            ),
        }

        if isinstance(
            exception,
            DataIntegrityError,
        ):
            payload["integrity_details"] = (
                exception.details
            )

    if outcome is not None:
        payload["outcome"] = {
            "status": outcome.status,
            "record_count": outcome.record_count,
            "message": outcome.message,
            "retry_later": outcome.retry_later,
            "duplicate_keys": outcome.duplicate_keys,
            "page_audits": [
                asdict(item)
                for item in outcome.page_audits
            ],
        }

    temp_path = path.with_suffix(".json.tmp")
    temp_path.write_text(
        json.dumps(
            json_safe(payload),
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    os.replace(temp_path, path)

    latest_path = error_log_dir / (
        f"{company.query_id}_latest.json"
    )

    try:
        shutil.copy2(path, latest_path)
    except OSError:
        pass

    return path


def is_captcha_related_exception(
    exception: BaseException,
) -> bool:
    """判斷例外是否發生在 CAPTCHA 擷取、辨識、輸入或驗證流程。"""
    if isinstance(
        exception,
        CaptchaRecognitionError,
    ):
        return True

    message = normalize_text(
        f"{type(exception).__name__}: {exception}"
    ).lower()

    captcha_keywords = (
        "captcha",
        "驗證碼",
        "imageinput",
        "ocrimage",
        "englishalphanumeric",
        "四個英數",
        "四碼",
    )

    return any(
        keyword in message
        for keyword in captcha_keywords
    )


def safe_page_url(
    page: Page | None,
) -> str:
    """在頁面已關閉或導頁中的情況下安全取得 URL。"""
    if page is None:
        return ""

    try:
        return str(page.url)
    except Exception:
        return ""


def page_is_open(
    page: Page | None,
) -> bool:
    """判斷 Playwright Page 是否仍可使用。"""
    if page is None:
        return False

    try:
        return not page.is_closed()
    except Exception:
        return False


def browser_is_connected(
    browser: Browser | None,
) -> bool:
    """判斷 Playwright Browser process 是否仍連線。"""
    if browser is None:
        return False

    try:
        return bool(browser.is_connected())
    except Exception:
        return False


def is_browser_session_closed_exception(
    exception: BaseException,
) -> bool:
    """
    判斷 Page／BrowserContext／Browser 是否已關閉。

    這類錯誤不能在原本的 page 上重試，必須先重建瀏覽器工作階段。
    """
    message = normalize_text(
        f"{type(exception).__name__}: {exception}"
    ).lower()

    closed_patterns = (
        "targetclosederror",
        "target page, context or browser has been closed",
        "page has been closed",
        "browser has been closed",
        "context has been closed",
        "browser disconnected",
        "browser.new_context: target page, context or browser has been closed",
    )

    return any(
        pattern in message
        for pattern in closed_patterns
    )


def click_effect_observed(
    page: Page,
    locator: Locator,
    before_url: str,
    before_text: str,
) -> bool:
    """
    確認點擊是否其實已經成功。

    Playwright 可能只是在「等待導頁完成」時逾時；只要 URL、DOM、
    結果標記或控制項附著狀態已改變，就視為點擊已成功觸發。
    """
    if not page_is_open(page):
        return False

    current_url = safe_page_url(page)

    if current_url and current_url != before_url:
        return True

    try:
        if locator.count() <= 0:
            return True
    except PlaywrightError:
        # navigation 造成 execution context／locator 失效時，通常表示點擊
        # 已觸發；但頁面若已真正關閉，前面已先返回 False。
        return page_is_open(page)

    if before_text:
        current_text = body_text(page)
        if current_text and current_text != before_text:
            return True

    return False


def wait_for_click_effect(
    page: Page,
    locator: Locator,
    before_url: str,
    before_text: str,
    timeout_ms: int,
) -> bool:
    """短暫等待 URL 或 DOM 改變，避免把成功導頁誤判成 click 失敗。"""
    deadline = time.monotonic() + max(
        250,
        min(timeout_ms, 5000),
    ) / 1000.0

    while time.monotonic() < deadline:
        if click_effect_observed(
            page,
            locator,
            before_url,
            before_text,
        ):
            return True
        time.sleep(0.10)

    return click_effect_observed(
        page,
        locator,
        before_url,
        before_text,
    )


def is_transient_page_exception(
    exception: BaseException,
) -> bool:
    """判斷是否為可重新查詢的暫時性頁面錯誤。"""
    if isinstance(
        exception,
        TransientPageError,
    ):
        return True

    if is_browser_session_closed_exception(
        exception
    ):
        return True

    message = normalize_text(
        f"{type(exception).__name__}: {exception}"
    ).lower()

    transient_patterns = (
        "execution context was destroyed",
        "cannot find context",
        "most likely because of a navigation",
        "navigation interrupted",
        "frame was detached",
        "page.goto: timeout",
        "net::err_",
    )

    return any(
        pattern in message
        for pattern in transient_patterns
    )

def wait_for_result_change(
    page: Page,
    before_text: str,
    timeout_ms: int,
) -> None:
    """等待一般頁面或分頁內容改變。"""
    try:
        page.wait_for_function(
            """
            before => {
              const now = (
                document.body?.innerText || ''
              )
                .replace(/\\s+/g, ' ')
                .trim();

              return now !== before;
            }
            """,
            arg=before_text,
            timeout=timeout_ms,
        )
    except PlaywrightTimeoutError:
        pass
    except PlaywrightError:
        # 導頁時 execution context 會被銷毀，後續以 load state 接手。
        pass

    try:
        page.wait_for_load_state(
            "domcontentloaded",
            timeout=min(timeout_ms, 15000),
        )
    except PlaywrightTimeoutError:
        pass
    except PlaywrightError:
        pass

    try:
        page.wait_for_load_state(
            "networkidle",
            timeout=min(timeout_ms, 10000),
        )
    except PlaywrightTimeoutError:
        pass
    except PlaywrightError:
        pass


def read_no_paid_data_message(
    page: Page,
) -> str:
    """
    讀取結果頁「查無已繳納罰鍰資料」訊息。

    第一優先使用使用者提供的精確 XPath；若網站只做小幅版面調整，
    再使用 id/class/文字語意備援。回傳空字串代表尚未確認為零筆結果。
    """
    candidates = (
        page.locator(
            f"xpath={NO_PAID_DATA_XPATH}"
        ),
        page.locator(
            "#headerMessage"
        ),
        page.get_by_text(
            re.compile(
                r"查無已繳納罰鍰資料",
                re.I,
            ),
            exact=False,
        ),
        page.locator(
            "td:has-text('查無已繳納罰鍰資料')"
        ),
    )

    for candidate in candidates:
        try:
            count = candidate.count()
        except PlaywrightError:
            continue

        for index in range(count):
            item = candidate.nth(index)

            try:
                if not item.is_visible():
                    continue

                message = normalize_text(
                    item.inner_text(
                        timeout=3000
                    )
                )
            except PlaywrightError:
                continue

            if contains_any(
                message,
                (
                    "查無已繳納罰鍰資料",
                    "無已繳納罰鍰資料",
                    "查無罰鍰繳納資料",
                ),
            ):
                return message

    return ""


def page_has_query_form(page: Page) -> bool:
    """判斷目前是否仍停留在統一編號／CAPTCHA 查詢頁。"""
    selectors = (
        "form#form2",
        f"xpath={QUERY_BUTTON_XPATH}",
        f"xpath={CURRENT_QUERY_BUTTON_XPATH}",
        "input[name*='captcha' i]",
        "input[id*='captcha' i]",
        "input[name*='verify' i]",
        "input[id*='verify' i]",
    )

    for selector in selectors:
        try:
            if page.locator(selector).count() > 0:
                return True
        except PlaywrightError:
            continue

    return False


def page_has_result_marker(page: Page) -> bool:
    """判斷結果頁的正式容器、表格或訊息是否已出現。"""
    selectors = (
        "form#penaltyQueryPayRecordForm",
        "table#info",
        "#headerMessage",
        "div.cont90 div.caption_std",
        f"xpath={RESULT_JSON_XPATH}",
        f"xpath={NO_PAID_DATA_XPATH}",
        f"xpath={CURRENT_RESULT_CONTAINER_XPATH}",
        f"xpath={CURRENT_RESULT_TABLE_XPATH}",
        f"xpath={CURRENT_RESULT_SECOND_TAB_XPATH}",
        f"xpath={CURRENT_PAGINATION_BANNER_XPATH}",
        f"xpath={PAID_PAGINATION_BANNER_XPATH}",
    )

    for selector in selectors:
        try:
            if page.locator(selector).count() > 0:
                return True
        except PlaywrightError:
            continue

    return False


def wait_for_query_result_page(
    page: Page,
    before_text: str,
    timeout_ms: int,
) -> str:
    """
    等待查詢結果真正穩定，而不是只等到第一次 DOM 文字變動。

    舊版在導頁剛開始時就繼續讀取，會出現：
    - 尚未顯示「驗證碼輸入錯誤」便開始解析；
    - Page.evaluate 的 execution context 被導頁銷毀；
    - 只取得 735 bytes 的暫時空白頁。
    """
    deadline = time.monotonic() + max(
        timeout_ms,
        1000,
    ) / 1000.0

    last_signature: tuple[str, str] | None = None
    stable_count = 0
    latest_text = ""

    while time.monotonic() < deadline:
        try:
            page.wait_for_load_state(
                "domcontentloaded",
                timeout=1000,
            )
        except (
            PlaywrightTimeoutError,
            PlaywrightError,
        ):
            pass

        latest_text = body_text(page)

        marker_ready = (
            page_has_result_marker(page)
            or bool(
                read_no_paid_data_message(page)
            )
            or contains_any(
                latest_text,
                CAPTCHA_ERROR_PATTERNS,
            )
            or contains_any(
                latest_text,
                FORM_ERROR_PATTERNS,
            )
            or contains_any(
                latest_text,
                NO_DATA_PATTERNS,
            )
        )

        changed = bool(
            latest_text
            and latest_text != before_text
        )

        if marker_ready and changed:
            signature = (
                page.url,
                latest_text,
            )

            if signature == last_signature:
                stable_count += 1
            else:
                last_signature = signature
                stable_count = 1

            if stable_count >= 2:
                return latest_text
        else:
            stable_count = 0
            last_signature = None

        time.sleep(0.25)

    return body_text(page)

def compact_header_text(value: Any) -> str:
    return re.sub(
        r"[\s:：、，,。．·()（）\[\]【】]",
        "",
        normalize_text(value),
    )


def header_index(
    headers: Sequence[str],
    aliases: Sequence[str],
) -> int | None:
    normalized_aliases = {
        compact_header_text(alias)
        for alias in aliases
    }

    for index, header in enumerate(
        headers
    ):
        normalized = compact_header_text(
            header
        )

        # 導覽列或整頁摘要通常非常長，
        # 不可把其中出現的關鍵字當成表頭。
        if (
            not normalized
            or len(normalized) > 16
        ):
            continue

        if normalized in normalized_aliases:
            return index

    return None


def identify_result_header(
    rows: Sequence[Sequence[str]],
) -> tuple[
    str,
    int,
    dict[str, int],
] | None:
    """
    僅接受真正的違規結果表頭。

    paid：
        繳費日期、單號、車號、事由、繳納方式、罰鍰

    unpaid：
        違規日、事由、罰鍰、應到案日

    必須分散在不同儲存格中，避免將網站導覽列、
    整頁文字摘要誤判成明細表格。
    """
    for row_index, row in enumerate(
        rows
    ):
        headers = [
            normalize_text(value)
            for value in row
        ]

        paid_map = {
            "date": header_index(
                headers,
                (
                    "繳費日期",
                    "繳納日期",
                ),
            ),
            "ticket": header_index(
                headers,
                (
                    "單號",
                    "舉發單號",
                ),
            ),
            "plate": header_index(
                headers,
                (
                    "車號",
                    "車牌",
                    "車牌號碼",
                ),
            ),
            "reason": header_index(
                headers,
                (
                    "事由",
                    "違規事由",
                ),
            ),
            "method": header_index(
                headers,
                (
                    "繳納方式",
                    "繳費方式",
                    "處理方式",
                ),
            ),
            "fine": header_index(
                headers,
                (
                    "罰鍰",
                    "罰款",
                    "金額",
                ),
            ),
        }

        if all(
            value is not None
            for value in paid_map.values()
        ):
            distinct = {
                int(value)
                for value in paid_map.values()
                if value is not None
            }

            if len(distinct) == len(
                paid_map
            ):
                return (
                    TABLE_KIND_PAID,
                    row_index,
                    {
                        key: int(value)
                        for key, value
                        in paid_map.items()
                        if value is not None
                    },
                )

        unpaid_map = {
            "date": header_index(
                headers,
                (
                    "違規日",
                    "違規日期",
                ),
            ),
            "reason": header_index(
                headers,
                (
                    "事由",
                    "違規事由",
                ),
            ),
            "fine": header_index(
                headers,
                (
                    "罰鍰",
                    "罰款",
                    "金額",
                ),
            ),
            "due": header_index(
                headers,
                (
                    "應到案日",
                    "應到案日期",
                    "應到案",
                ),
            ),
        }

        if all(
            value is not None
            for value in unpaid_map.values()
        ):
            distinct = {
                int(value)
                for value in unpaid_map.values()
                if value is not None
            }

            if len(distinct) == len(
                unpaid_map
            ):
                return (
                    TABLE_KIND_UNPAID,
                    row_index,
                    {
                        key: int(value)
                        for key, value
                        in unpaid_map.items()
                        if value is not None
                    },
                )

    return None


def row_value(
    row: Sequence[str],
    index: int | None,
) -> str:
    if (
        index is None
        or index < 0
        or index >= len(row)
    ):
        return ""

    return normalize_text(
        row[index]
    )


def normalize_amount(value: Any) -> int | str:
    text = normalize_text(value)

    if not text:
        return ""

    compact = re.sub(
        r"[,，\s元]",
        "",
        text,
    )

    if re.fullmatch(
        r"-?\d+",
        compact,
    ):
        try:
            return int(compact)
        except ValueError:
            pass

    return text


def is_roc_date_text(value: Any) -> bool:
    text = normalize_text(value)

    return bool(
        re.fullmatch(
            r"\d{2,3}\s*年\s*"
            r"\d{1,2}\s*月\s*"
            r"\d{1,2}\s*日",
            text,
        )
    )


def shifted_row_value(
    row: Sequence[str],
    index: int,
    offset: int,
) -> str:
    return row_value(
        row,
        index + offset,
    )


def select_paid_column_offset(
    row: Sequence[str],
    columns: dict[str, int],
) -> int:
    best_offset = 0
    best_score = -1

    for offset in (
        -1,
        0,
        1,
        2,
    ):
        payment_date = shifted_row_value(
            row,
            columns["date"],
            offset,
        )

        ticket_number = shifted_row_value(
            row,
            columns["ticket"],
            offset,
        )

        plate_number = shifted_row_value(
            row,
            columns["plate"],
            offset,
        )

        reason = shifted_row_value(
            row,
            columns["reason"],
            offset,
        )

        amount = normalize_amount(
            shifted_row_value(
                row,
                columns["fine"],
                offset,
            )
        )

        score = 0

        if (
            not payment_date
            or is_roc_date_text(
                payment_date
            )
        ):
            score += 1

        if re.fullmatch(
            r"[A-Za-z0-9-]{5,24}",
            ticket_number,
        ):
            score += 3

        if (
            plate_number
            and len(plate_number) <= 20
        ):
            score += 2

        if reason:
            score += 2

        if amount not in ("", None):
            score += 2

        if score > best_score:
            best_score = score
            best_offset = offset

    return best_offset


def select_unpaid_column_offset(
    row: Sequence[str],
    columns: dict[str, int],
) -> int:
    best_offset = 0
    best_score = -1

    for offset in (
        -1,
        0,
        1,
        2,
    ):
        violation_date = shifted_row_value(
            row,
            columns["date"],
            offset,
        )

        reason = shifted_row_value(
            row,
            columns["reason"],
            offset,
        )

        amount = normalize_amount(
            shifted_row_value(
                row,
                columns["fine"],
                offset,
            )
        )

        due_date = shifted_row_value(
            row,
            columns["due"],
            offset,
        )

        score = 0

        if is_roc_date_text(
            violation_date
        ):
            score += 4

        if reason:
            score += 2

        if amount not in ("", None):
            score += 2

        if (
            not due_date
            or is_roc_date_text(
                due_date
            )
        ):
            score += 2

        if score > best_score:
            best_score = score
            best_offset = offset

    return best_offset


def canonicalize_paid_row(
    row: Sequence[str],
    columns: dict[str, int],
) -> list[Any] | None:
    offset = select_paid_column_offset(
        row,
        columns,
    )

    payment_date = shifted_row_value(
        row,
        columns["date"],
        offset,
    )

    ticket_number = shifted_row_value(
        row,
        columns["ticket"],
        offset,
    )

    plate_number = shifted_row_value(
        row,
        columns["plate"],
        offset,
    )

    reason = shifted_row_value(
        row,
        columns["reason"],
        offset,
    )

    payment_method = shifted_row_value(
        row,
        columns["method"],
        offset,
    )

    fine = normalize_amount(
        shifted_row_value(
            row,
            columns["fine"],
            offset,
        )
    )

    if (
        compact_header_text(payment_date)
        in {
            "繳費日期",
            "繳納日期",
        }
    ):
        return None

    # 來源檔中有少數繳費日期空白的紀錄，
    # 因此以單號、車號、事由及罰鍰判斷資料列。
    if not (
        ticket_number
        and plate_number
        and reason
        and fine not in ("", None)
    ):
        return None

    if (
        payment_date
        and not is_roc_date_text(
            payment_date
        )
    ):
        return None

    return [
        None,
        payment_date or None,
        ticket_number,
        plate_number,
        reason,
        payment_method or None,
        fine,
    ]


def canonicalize_unpaid_row(
    row: Sequence[str],
    columns: dict[str, int],
) -> list[Any] | None:
    offset = select_unpaid_column_offset(
        row,
        columns,
    )

    violation_date = shifted_row_value(
        row,
        columns["date"],
        offset,
    )

    reason = shifted_row_value(
        row,
        columns["reason"],
        offset,
    )

    fine = normalize_amount(
        shifted_row_value(
            row,
            columns["fine"],
            offset,
        )
    )

    due_date = shifted_row_value(
        row,
        columns["due"],
        offset,
    )

    if (
        compact_header_text(
            violation_date
        )
        in {
            "違規日",
            "違規日期",
        }
    ):
        return None

    if not (
        is_roc_date_text(
            violation_date
        )
        and reason
        and fine not in ("", None)
    ):
        return None

    if (
        due_date
        and not is_roc_date_text(
            due_date
        )
    ):
        return None

    combined = " ".join(
        normalize_text(value)
        for value in row
    )

    status = (
        "需到案"
        if "需到案" in combined
        else None
    )

    view_text = next(
        (
            normalize_text(value)
            for value in row
            if "檢視" in normalize_text(
                value
            )
        ),
        "檢視",
    )

    return [
        status,
        violation_date,
        reason,
        fine,
        due_date or None,
        view_text,
        None,
    ]

ENGLISH_MONTH_NUMBERS = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def parse_result_datetime(
    value: Any,
) -> datetime | None:
    """
    解析結果 JSON 常見日期格式。

    支援例如：
    - Jul 2, 2026 8:50:26 AM
    - Thu Jul 02 08:50:26 CST 2026
    - 2026/07/02 08:50:26
    - 2026-07-02 08:50:26
    """
    if isinstance(value, datetime):
        return value

    if (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
    ):
        timestamp = float(value)

        if abs(timestamp) >= 10_000_000_000:
            timestamp /= 1000.0

        try:
            return datetime.fromtimestamp(
                timestamp
            )
        except (
            OverflowError,
            OSError,
            ValueError,
        ):
            return None

    text = normalize_text(value)

    if not text:
        return None

    normalized = re.sub(
        r"\s+",
        " ",
        text,
    ).strip()

    for date_format in (
        "%b %d, %Y %I:%M:%S %p",
        "%b %d, %Y %I:%M %p",
        "%b %d, %Y",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%a %b %d %H:%M:%S %Z %Y",
    ):
        try:
            return datetime.strptime(
                normalized,
                date_format,
            )
        except ValueError:
            continue

    match = re.search(
        r"(?P<month>[A-Za-z]{3,9})\s+"
        r"(?P<day>\d{1,2})(?:,)?\s+"
        r"(?P<year>\d{4})",
        normalized,
    )

    if match:
        month = ENGLISH_MONTH_NUMBERS.get(
            match.group("month")[:3].lower()
        )

        if month is not None:
            try:
                return datetime(
                    int(match.group("year")),
                    month,
                    int(match.group("day")),
                )
            except ValueError:
                return None

    match = re.search(
        r"(?:[A-Za-z]{3}\s+)?"
        r"(?P<month>[A-Za-z]{3,9})\s+"
        r"(?P<day>\d{1,2})\s+"
        r"(?:\d{1,2}:\d{2}(?::\d{2})?\s+)?"
        r"(?:[A-Za-z]{2,5}\s+)?"
        r"(?P<year>\d{4})",
        normalized,
    )

    if match:
        month = ENGLISH_MONTH_NUMBERS.get(
            match.group("month")[:3].lower()
        )

        if month is not None:
            try:
                return datetime(
                    int(match.group("year")),
                    month,
                    int(match.group("day")),
                )
            except ValueError:
                return None

    return None


def format_roc_date(
    value: Any,
) -> str:
    """
    將西元日期轉為來源 Excel 使用的民國日期文字。

    例如：
        Jul 2, 2026 8:50:26 AM
        -> 115年7月2日
    """
    parsed = parse_result_datetime(
        value
    )

    if parsed is None:
        return normalize_text(
            value
        )

    return (
        f"{parsed.year - 1911}年"
        f"{parsed.month}月"
        f"{parsed.day}日"
    )


def first_present_value(
    record: dict[str, Any],
    keys: Sequence[str],
) -> Any:
    for key in keys:
        value = record.get(key)

        if value is None:
            continue

        if isinstance(value, str):
            if normalize_text(value):
                return value
            continue

        return value

    return None


def build_paid_table_from_json_payload(
    payload: Any,
) -> ExtractedTable | None:
    """將目前頁面的 hidden JSON 轉成來源 155598 的 A:G 結構。"""
    records: Any = payload

    if isinstance(records, dict):
        for key in (
            "data",
            "records",
            "items",
            "rows",
            "result",
        ):
            candidate = records.get(key)

            if isinstance(candidate, list):
                records = candidate
                break

    if not isinstance(records, list):
        return None

    canonical_rows: list[list[Any]] = []
    row_keys: list[str] = []

    for raw_record in records:
        if not isinstance(raw_record, dict):
            continue

        payment_date_value = first_present_value(
            raw_record,
            (
                "updateTimeStr",
                "updateTime",
                "paymentDateStr",
                "paymentDate",
                "payDateStr",
                "payDate",
            ),
        )

        ticket_number = normalize_text(
            first_present_value(
                raw_record,
                (
                    "vilTicket",
                    "ticket",
                    "ticketNo",
                ),
            )
        )
        plate_number = normalize_text(
            first_present_value(
                raw_record,
                (
                    "plateNo",
                    "plate",
                    "plateNumber",
                ),
            )
        )
        reason = normalize_text(
            first_present_value(
                raw_record,
                (
                    "vilFact",
                    "reason",
                    "fact",
                ),
            )
        )
        payment_method = normalize_text(
            first_present_value(
                raw_record,
                (
                    "payWay",
                    "paymentMethod",
                    "payMethod",
                ),
            )
        )
        fine = normalize_amount(
            first_present_value(
                raw_record,
                (
                    "payment",
                    "penalty",
                    "penaltyAmount",
                    "amount",
                ),
            )
        )

        if not (
            ticket_number
            and plate_number
            and reason
            and fine not in ("", None)
        ):
            continue

        payment_date = format_roc_date(
            payment_date_value
        )
        canonical = [
            None,
            payment_date or None,
            ticket_number,
            plate_number,
            reason,
            payment_method or None,
            fine,
        ]
        canonical_rows.append(canonical)

        # 單號是已繳紀錄的主要識別欄位；仍將其他欄位納入，避免網站重用單號時誤判。
        row_keys.append(
            record_fingerprint(
                [
                    ticket_number,
                    plate_number,
                    payment_date,
                    reason,
                    payment_method,
                    fine,
                ]
            )
        )

    if not canonical_rows:
        return None

    return ExtractedTable(
        kind=TABLE_KIND_PAID,
        title="罰鍰繳納紀錄",
        headers=[
            "",
            "繳費日期",
            "單號",
            "車號",
            "事由",
            "繳納方式",
            "罰鍰",
        ],
        rows=canonical_rows,
        row_keys=row_keys,
    )

def extract_paid_table_from_hidden_json(
    page: Page,
) -> tuple[bool, ExtractedTable | None]:
    """
    從結果頁 hidden input 讀取完整 JSON。

    回傳：
        (是否找到 JSON input, 解析後的繳納紀錄表)

    找到 input 但 value 為 [] 時，回傳 (True, None)。
    """
    candidates = [
        page.locator(
            f"xpath={RESULT_JSON_XPATH}"
        ),
        page.locator(
            "form#penaltyQueryPayRecordForm input#json"
        ),
        page.locator(
            "form#penaltyQueryPayRecordForm input[name='json']"
        ),
        page.locator(
            "input#json[name='json']"
        ),
    ]

    found_input = False
    last_error: Exception | None = None

    for candidate in candidates:
        try:
            count = candidate.count()
        except PlaywrightError as exc:
            last_error = exc
            continue

        if count <= 0:
            continue

        for index in range(count):
            item = candidate.nth(index)

            try:
                item.wait_for(
                    state="attached",
                    timeout=5000,
                )

                element_id = normalize_text(
                    item.get_attribute("id")
                )

                element_name = normalize_text(
                    item.get_attribute("name")
                )

                if (
                    element_id
                    and element_id != "json"
                    and element_name != "json"
                ):
                    continue

                raw_value = item.get_attribute(
                    "value"
                )

                found_input = True

                if raw_value is None:
                    continue

                stripped = raw_value.strip()

                if not stripped:
                    return True, None

                try:
                    payload = json.loads(
                        stripped
                    )
                except json.JSONDecodeError:
                    # 若取得的是 HTML 原始屬性文字，
                    # 將 &quot; 等 entity 還原後再解析。
                    payload = json.loads(
                        unescape(
                            stripped
                        )
                    )

                return (
                    True,
                    build_paid_table_from_json_payload(
                        payload
                    ),
                )

            except (
                json.JSONDecodeError,
                PlaywrightError,
                TypeError,
                ValueError,
            ) as exc:
                last_error = exc
                continue

    if (
        found_input
        and last_error is not None
    ):
        raise RuntimeError(
            "已找到結果 JSON input，"
            "但無法解析其 value："
            f"{type(last_error).__name__}: "
            f"{last_error}"
        ) from last_error

    return False, None

def extract_visible_tables(
    page: Page,
) -> list[ExtractedTable]:
    """
    從頁面中擷取真正的結果表格。

    原版本只要表格文字含有「交通違規、罰鍰、車牌」等字樣，
    就可能把網站導覽選單當成結果，造成輸出出現 A:P 的巨大雜訊。

    本版本要求表格具有完整且分欄的正式表頭，
    並將資料直接轉成來源 155598 的 A:G 欄位結構。
    """
    payload = page.evaluate(
        """
        () => {
          const norm = value => (
            value || ''
          )
            .replace(/\\s+/g, ' ')
            .trim();

          const visible = element => {
            const style = getComputedStyle(
              element
            );

            const rect = (
              element
              .getBoundingClientRect()
            );

            return (
              style.visibility !== 'hidden'
              && style.display !== 'none'
              && rect.width > 0
              && rect.height > 0
            );
          };

          const directRows = table => {
            const selectors = [
              ':scope > thead > tr',
              ':scope > tbody > tr',
              ':scope > tfoot > tr',
              ':scope > tr'
            ].join(', ');

            return [
              ...table.querySelectorAll(
                selectors
              )
            ].filter(visible);
          };

          const directCells = row => [
            ...row.children
          ]
            .filter(cell => (
              (
                cell.tagName === 'TH'
                || cell.tagName === 'TD'
              )
              && visible(cell)
            ))
            .map(cell => {
              const controls = [
                ...cell.querySelectorAll(
                  'input, a, button, select, option'
                )
              ].map(control => [
                control.tagName || '',
                control.getAttribute('type') || '',
                control.getAttribute('name') || '',
                control.id || '',
                control.getAttribute('value') || '',
                control.getAttribute('href') || '',
                control.getAttribute('onclick') || '',
                control.getAttribute('data-id') || '',
                control.getAttribute('data-key') || ''
              ].join('|'));

              return {
                text: norm(cell.innerText),
                identity: norm([
                  cell.innerText || '',
                  cell.getAttribute('id') || '',
                  cell.getAttribute('data-id') || '',
                  cell.getAttribute('data-key') || '',
                  ...controls
                ].join(' || ')),
                header: cell.tagName === 'TH'
              };
            });

          const previousTitle = table => {
            if (
              table.caption
              && norm(
                table.caption.innerText
              )
            ) {
              return norm(
                table.caption.innerText
              );
            }

            let current = table;

            for (
              let depth = 0;
              depth < 6 && current;
              depth++
            ) {
              let previous = (
                current.previousElementSibling
              );

              while (previous) {
                if (
                  visible(previous)
                  && norm(
                    previous.innerText
                  )
                  && (
                    /^H[1-6]$/.test(
                      previous.tagName
                    )
                    || /(title|header|caption)/i
                      .test(
                        String(
                          previous.className
                          || ''
                        )
                      )
                  )
                ) {
                  return norm(
                    previous.innerText
                  );
                }

                previous = (
                  previous
                  .previousElementSibling
                );
              }

              current = (
                current.parentElement
              );
            }

            return '';
          };

          return [
            ...document.querySelectorAll(
              'table'
            )
          ]
            .filter(visible)
            .map(table => ({
              title: previousTitle(
                table
              ),
              rows: directRows(
                table
              )
                .map(directCells)
                .filter(row => (
                  row.length > 0
                ))
            }))
            .filter(table => (
              table.rows.length > 0
            ));
        }
        """
    )

    extracted: list[
        ExtractedTable
    ] = []

    seen: set[str] = set()

    for raw_table in payload:
        raw_rows = (
            raw_table.get("rows")
            or []
        )

        plain_rows = [
            [
                normalize_text(
                    cell.get("text")
                )
                for cell in row
            ]
            for row in raw_rows
        ]

        identified = identify_result_header(
            plain_rows
        )

        if identified is None:
            continue

        kind, header_row, columns = (
            identified
        )

        canonical_rows: list[
            list[Any]
        ] = []
        canonical_keys: list[str] = []

        for row_offset, row in enumerate(
            plain_rows[header_row + 1:]
        ):
            if kind == TABLE_KIND_PAID:
                canonical = (
                    canonicalize_paid_row(
                        row,
                        columns,
                    )
                )
            else:
                canonical = (
                    canonicalize_unpaid_row(
                        row,
                        columns,
                    )
                )

            if canonical is not None:
                canonical_rows.append(
                    canonical
                )
                raw_identity_row = [
                    normalize_text(
                        cell.get("identity")
                        or cell.get("text")
                    )
                    for cell in raw_rows[
                        header_row + 1 + row_offset
                    ]
                ]
                canonical_keys.append(
                    record_fingerprint(
                        raw_identity_row
                    )
                )

        if not canonical_rows:
            continue

        headers = plain_rows[
            header_row
        ]

        signature = hashlib.sha256(
            json.dumps(
                [
                    kind,
                    headers,
                    canonical_rows,
                ],
                ensure_ascii=False,
                sort_keys=False,
                default=str,
            ).encode("utf-8")
        ).hexdigest()

        if signature in seen:
            continue

        seen.add(
            signature
        )

        extracted.append(
            ExtractedTable(
                kind=kind,
                title=normalize_text(
                    raw_table.get(
                        "title"
                    )
                ),
                headers=headers,
                rows=canonical_rows,
                row_keys=canonical_keys,
            )
        )

    return extracted

def extract_visible_tables_with_retry(
    page: Page,
    timeout_ms: int,
) -> list[ExtractedTable]:
    """在導頁 execution context 切換期間重試 table extraction。"""
    deadline = time.monotonic() + min(
        max(timeout_ms, 1000),
        15000,
    ) / 1000.0

    last_error: Exception | None = None

    while time.monotonic() < deadline:
        try:
            return extract_visible_tables(page)
        except PlaywrightError as exc:
            last_error = exc

            if not is_transient_page_exception(exc):
                raise

            try:
                page.wait_for_load_state(
                    "domcontentloaded",
                    timeout=1000,
                )
            except (
                PlaywrightTimeoutError,
                PlaywrightError,
            ):
                pass

            time.sleep(0.25)

    if last_error is not None:
        raise TransientPageError(
            "結果頁在導向期間無法穩定讀取表格："
            f"{type(last_error).__name__}: {last_error}"
        ) from last_error

    return []

def is_record_row(
    row: Sequence[str],
) -> bool:
    text = " ".join(
        normalize_text(value)
        for value in row
    )

    if not text:
        return False

    if contains_any(
        text,
        NO_DATA_PATTERNS,
    ):
        return False

    if re.search(
        r"(^|\s)"
        r"(合計|總計|共計|總筆數|資料筆數)"
        r"(\s|$)",
        text,
    ):
        return False

    if sum(
        keyword in text
        for keyword in RESULT_KEYWORDS
    ) >= 2:
        return False

    if re.search(
        r"\d{2,3}\s*年\s*"
        r"\d{1,2}\s*月\s*"
        r"\d{1,2}\s*日",
        text,
    ):
        return True

    if (
        re.search(
            r"\b[A-Z0-9]{6,12}\b",
            text,
        )
        and len(
            [value for value in row if value]
        ) >= 3
    ):
        return True

    if (
        re.search(
            r"(?:^|\s)"
            r"\d{2,6}"
            r"(?:元)?"
            r"(?:\s|$)",
            text,
        )
        and len(
            [value for value in row if value]
        ) >= 3
    ):
        return True

    return False


def find_next_page_control(
    page: Page,
) -> Locator | None:
    candidates = [
        page.get_by_role(
            "link",
            name=re.compile(
                r"^(下一頁|下頁|Next|›|»)$",
                re.I,
            ),
        ),
        page.get_by_role(
            "button",
            name=re.compile(
                r"^(下一頁|下頁|Next|›|»)$",
                re.I,
            ),
        ),
        page.locator(
            "input[type='button']"
            "[value='下一頁']"
        ),
        page.locator(
            "input[type='submit']"
            "[value='下一頁']"
        ),
    ]

    for candidate in candidates:
        try:
            for index in range(
                candidate.count()
            ):
                item = candidate.nth(index)

                if (
                    not item.is_visible()
                    or not item.is_enabled()
                ):
                    continue

                disabled = (
                    item.get_attribute(
                        "aria-disabled"
                    )
                    == "true"
                )

                classes = (
                    item.get_attribute("class")
                    or ""
                ).lower()

                if (
                    disabled
                    or "disabled" in classes
                ):
                    continue

                return item
        except PlaywrightError:
            continue

    return None


def extract_paid_page_tables(
    page: Page,
    timeout_ms: int,
) -> list[ExtractedTable]:
    """
    擷取第二階段目前頁面的資料。

    優先使用畫面上目前頁的表格，因為 hidden JSON 不保證包含全部頁，
    也不保證能正確反映目前頁碼。只有畫面表格無法解析時，才使用
    hidden JSON 作為同一頁的備援。
    """
    visible_tables = extract_visible_tables_with_retry(
        page,
        timeout_ms,
    )
    paid_visible = [
        table
        for table in visible_tables
        if table.kind == TABLE_KIND_PAID
    ]

    if paid_visible:
        return paid_visible

    json_input_found, paid_json_table = (
        extract_paid_table_from_hidden_json(page)
    )

    if json_input_found and paid_json_table is not None:
        return [paid_json_table]

    return []


def collect_all_result_pages(
    page: Page,
    timeout_ms: int,
    max_pages: int = 1000,
) -> tuple[list[ExtractedTable], list[PageAudit]]:
    """
    逐頁收集「交通違規繳納記錄查詢」。

    不再假設 hidden JSON 是全部資料；每一頁均以 goPage/changePage
    實際切換、擷取並驗證，最後必須符合網站 showbanner 宣告的
    「總頁數／總筆數」，且不得有重複資料。
    """
    return capture_paginated_group(
        page=page,
        banner_xpath=PAID_PAGINATION_BANNER_XPATH,
        section=QUERY_SECTION_PAID,
        category=PAID_RECORD_CATEGORY,
        extractor=lambda: extract_paid_page_tables(
            page,
            timeout_ms,
        ),
        timeout_ms=timeout_ms,
        max_pages=max_pages,
    )

def save_debug_artifacts(
    page: Page,
    debug_dir: Path,
    company_id: str,
    label: str,
) -> None:
    debug_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    stamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    safe_label = re.sub(
        r"[^A-Za-z0-9_-]",
        "_",
        label,
    )

    base = debug_dir / (
        f"{company_id}_"
        f"{stamp}_"
        f"{safe_label}"
    )

    try:
        base.with_suffix(".html").write_text(
            page.content(),
            encoding="utf-8",
        )
    except PlaywrightError:
        pass

    try:
        page.screenshot(
            path=str(
                base.with_suffix(".png")
            ),
            full_page=True,
        )
    except PlaywrightError:
        pass


def normalize_captcha_image_format(
    value: str,
) -> str:
    normalized = normalize_text(value).lower()

    if normalized == "jpeg":
        normalized = "jpg"

    if normalized not in {"png", "jpg"}:
        raise ValueError(
            "驗證碼圖片格式只支援 "
            "png、jpg 或 jpeg"
        )

    return normalized


def save_captcha_image(
    page: Page,
    captcha_dir: Path,
    company: CompanyRow,
    image_format: str,
) -> Path:
    """
    優先使用指定 XPath 擷取目前瀏覽器中的 CAPTCHA 圖片。

    回傳的 PNG/JPG 路徑會交給
    call_english_alphanumeric_ocr()，
    再由該函式呼叫 ocrImage(image)。
    """
    normalized_format = (
        normalize_captcha_image_format(
            image_format
        )
    )

    captcha_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    captcha_capture_area = locate_captcha_capture_area(page)

    captcha_capture_area.wait_for(
        state="visible",
        timeout=10000,
    )

    captcha_capture_area.scroll_into_view_if_needed(
        timeout=5000
    )

    stamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S_%f"
    )

    archive_path = captcha_dir / (
        f"{company.query_id}_"
        f"{stamp}_"
        f"captcha.{normalized_format}"
    )

    screenshot_arguments: dict[str, Any] = {
        "path": str(archive_path),
        "timeout": 10000,
        "type": (
            "png"
            if normalized_format == "png"
            else "jpeg"
        ),
    }

    if normalized_format == "jpg":
        screenshot_arguments["quality"] = 95

    captcha_capture_area.screenshot(
        **screenshot_arguments
    )

    if (
        not archive_path.exists()
        or archive_path.stat().st_size == 0
    ):
        raise RuntimeError(
            "驗證碼圖片截圖失敗或檔案為空"
        )

    latest_path = captcha_dir / (
        f"current_captcha."
        f"{normalized_format}"
    )

    try:
        shutil.copy2(
            archive_path,
            latest_path,
        )
    except OSError:
        pass

    return archive_path



def validate_captcha_code(value: Any) -> str:
    """
    驗證 OCR 回傳值。格式不符屬於可重試的 CAPTCHA 辨識失敗，
    不應直接跳出單筆查詢並寫成「錯誤」。
    """
    if not isinstance(value, str):
        raise CaptchaRecognitionError(
            "ocrImage() 必須回傳 str；"
            f"目前回傳型別：{type(value).__name__}"
        )

    code = normalize_text(value)

    if not re.fullmatch(r"[A-Za-z0-9]{4}", code):
        raise CaptchaRecognitionError(
            "ocrImage() 回傳值格式錯誤；"
            "必須恰好是四個英數字元，"
            f"目前收到：{code!r}"
        )

    return code

def call_english_alphanumeric_ocr(
    captcha_image_path: Path,
) -> str:
    """
    將 CAPTCHA 圖片載入為 PIL.Image.Image，直接呼叫：

        ocrImage(image)

    englishAlphanumericOcrApi.py 內的 ocrImage() 應接收
    PIL.Image.Image 並回傳四個英數字元字串。

    本函式接收 ocrImage() 的回傳值，驗證格式後再回傳給
    acquire_captcha()。
    """
    resolved_image_path = (
        captcha_image_path
        .expanduser()
        .resolve()
    )

    if not resolved_image_path.exists():
        raise RuntimeError(
            "要傳給 ocrImage() 的圖片不存在："
            f"{resolved_image_path}"
        )

    if not callable(ocrImage):
        raise RuntimeError(
            "從 englishAlphanumericOcrApi 匯入的 ocrImage 不是可呼叫函式"
        )

    ui_image: Image.Image | None = None

    print(
        "  呼叫 ocrImage(image)",
        flush=True,
    )
    print(
        f"  傳入 CAPTCHA 圖片：{resolved_image_path}",
        flush=True,
    )

    try:
        with Image.open(resolved_image_path) as opened_image:
            opened_image.load()

            # copy() 後，即使 Image.open() 的檔案已關閉，
            # ocrImage() 仍可正常使用圖片。
            ui_image = opened_image.copy()

        returned_code = ocrImage(
            ui_image
        )

    except Exception as exc:
        raise RuntimeError(
            "ocrImage() 執行失敗："
            f"{type(exc).__name__}: {exc}"
        ) from exc

    finally:
        if ui_image is not None:
            try:
                ui_image.close()
            except Exception:
                pass

    captcha_code = validate_captcha_code(
        returned_code
    )

    print(
        "  ocrImage() 已回傳四個英數字元，"
        "準備自動填入網頁驗證碼欄位。",
        flush=True,
    )

    return captcha_code


def acquire_captcha(
    page: Page,
    company: CompanyRow,
    captcha_dir: Path,
    captcha_image_format: str,
) -> str:
    """
    完整 CAPTCHA 人工輸入流程：

    1. 找到網頁驗證碼輸入欄位。
    2. 依 XPath 擷取目前 CAPTCHA 所在的 td 區塊。
    3. 保存成 PNG/JPG。
    4. 將圖片開啟成 PIL.Image.Image。
    5. 呼叫 ocrImage(image)。
    6. 接收 ocrImage() 回傳的四碼字串。
    7. 將字串回傳給 perform_query() 自動填入網頁。
    """
    captcha_input = locate_captcha_input(page)
    captcha_input.click()

    captcha_path = save_captcha_image(
        page=page,
        captcha_dir=captcha_dir,
        company=company,
        image_format=captcha_image_format,
    )

    normalized_format = (
        normalize_captcha_image_format(
            captcha_image_format
        )
    )

    print(
        f"  CAPTCHA 圖片已提取：{captcha_path}",
        flush=True,
    )

    print(
        "  最新驗證碼圖片固定路徑："
        f"{captcha_dir / ('current_captcha.' + normalized_format)}",
        flush=True,
    )

    return call_english_alphanumeric_ocr(
        captcha_image_path=captcha_path,
    )

def is_login_page(page: Page) -> bool:
    text = body_text(page)

    try:
        has_password = (
            page.locator(
                "input[type='password']"
            ).count()
            > 0
        )
    except PlaywrightError:
        has_password = False

    return (
        has_password
        or contains_any(
            text,
            (
                "會員登入",
                "登入監理服務網",
                "監理服務網會員",
            ),
        )
    )


def ensure_query_form_after_optional_login(
    page: Page,
    query_url: str,
    timeout_ms: int,
    headless: bool,
) -> Locator:
    try:
        return locate_company_id_input(page)
    except RuntimeError as original_error:
        if not is_login_page(page):
            raise

        if headless:
            raise RuntimeError(
                "網站要求會員登入，"
                "但目前使用 --headless；"
                "請移除 --headless 後重新執行"
            ) from original_error

        print(
            "網站目前要求會員登入。"
            "請在瀏覽器完成登入，"
            "完成後回到此視窗按 Enter。",
            flush=True,
        )

        try:
            input()
        except EOFError as exc:
            raise RuntimeError(
                "無法等待人工登入"
            ) from exc

        page.goto(
            query_url,
            wait_until="domcontentloaded",
            timeout=timeout_ms,
        )

        try:
            page.wait_for_load_state(
                "networkidle",
                timeout=min(
                    timeout_ms,
                    15000,
                ),
            )
        except PlaywrightTimeoutError:
            pass

        return locate_company_id_input(page)



def click_locator_with_fallback(
    page: Page,
    locator: Locator,
    exact_xpath: str,
    timeout_ms: int,
    label: str,
) -> str:
    """
    點擊控制項，但不讓 Playwright 自動等待導頁拖垮整個流程。

    修正重點：
    1. 使用 no_wait_after=True，點擊後由呼叫端自行等待結果頁。
    2. 即使 click() 丟出 TimeoutError，也先檢查 URL／DOM 是否已改變。
    3. 已成功觸發導頁時不會再點第二次舊 locator。
    4. JavaScript 備援使用 setTimeout(element.click, 0)，讓 evaluate 在
       導頁發生前先返回，避免 execution context destroyed。
    """
    if not page_is_open(page):
        raise RuntimeError(
            f"{label}無法點擊：目前 Page 已關閉"
        )

    before_url = safe_page_url(page)
    before_text = body_text(page)
    errors: list[str] = []

    try:
        locator.scroll_into_view_if_needed(
            timeout=min(timeout_ms, 5000)
        )
    except PlaywrightError:
        pass

    try:
        locator.click(
            timeout=min(timeout_ms, 10000),
            no_wait_after=True,
        )
        return f"{label}一般點擊（不等待導頁）"
    except (TypeError, PlaywrightError) as exc:
        errors.append(
            f"一般點擊：{type(exc).__name__}: {exc}"
        )

        if wait_for_click_effect(
            page,
            locator,
            before_url,
            before_text,
            min(timeout_ms, 3000),
        ):
            return (
                f"{label}一般點擊已觸發；"
                "僅 Playwright 等待導頁逾時"
            )

    if is_browser_session_closed_exception(
        RuntimeError(" | ".join(errors))
    ) or not page_is_open(page):
        raise RuntimeError(
            f"{label}點擊後瀏覽器工作階段已關閉；"
            f"錯誤：{' | '.join(errors)}"
        )

    # 先用非同步 JavaScript 觸發，確保 evaluate 能在導頁前返回。
    try:
        clicked = page.evaluate(
            """
            xpath => {
              const element = document.evaluate(
                xpath,
                document,
                null,
                XPathResult.FIRST_ORDERED_NODE_TYPE,
                null
              ).singleNodeValue;

              if (!element) {
                return false;
              }

              window.setTimeout(() => element.click(), 0);
              return true;
            }
            """,
            exact_xpath,
        )

        if clicked:
            return f"{label}非同步 JavaScript 點擊"
    except PlaywrightError as exc:
        errors.append(
            "非同步 JavaScript 點擊："
            f"{type(exc).__name__}: {exc}"
        )

        if wait_for_click_effect(
            page,
            locator,
            before_url,
            before_text,
            min(timeout_ms, 3000),
        ):
            return (
                f"{label}JavaScript 點擊已觸發；"
                "導頁造成 execution context 切換"
            )

    if not page_is_open(page):
        raise RuntimeError(
            f"{label}點擊期間 Page 已關閉；"
            f"錯誤：{' | '.join(errors)}"
        )

    try:
        # 重新依 XPath 取得新 locator，避免沿用導頁前的舊 locator。
        fresh_locator = page.locator(
            f"xpath={exact_xpath}"
        ).first
        fresh_locator.click(
            timeout=min(timeout_ms, 10000),
            force=True,
            no_wait_after=True,
        )
        return f"{label}強制點擊（不等待導頁）"
    except (TypeError, PlaywrightError) as exc:
        errors.append(
            f"強制點擊：{type(exc).__name__}: {exc}"
        )

        try:
            fresh_locator = page.locator(
                f"xpath={exact_xpath}"
            ).first
        except Exception:
            fresh_locator = locator

        if wait_for_click_effect(
            page,
            fresh_locator,
            before_url,
            before_text,
            min(timeout_ms, 3000),
        ):
            return (
                f"{label}強制點擊已觸發；"
                "僅等待導頁逾時"
            )

    raise RuntimeError(
        f"{label}已定位，但所有點擊方式均失敗；"
        f"XPath：{exact_xpath}；"
        f"目前 URL：{safe_page_url(page)}；"
        f"錯誤：{' | '.join(errors)}"
    )

def locate_current_query_corporate_tab(
    page: Page,
) -> Locator | None:
    """定位第一階段查詢頁的「法人」頁籤。"""
    exact = page.locator(
        f"xpath={CURRENT_QUERY_CORPORATE_TAB_XPATH}"
    )

    try:
        if exact.count() > 0:
            return exact.first
    except PlaywrightError:
        pass

    candidates = (
        page.get_by_role(
            "img",
            name=re.compile(r"法人", re.I),
        ),
        page.locator(
            "img[alt*='法人']"
        ),
        page.locator(
            "img[title*='法人']"
        ),
        page.locator(
            "td:has-text('法人') span img"
        ),
    )

    for candidate in candidates:
        try:
            if candidate.count() > 0:
                return candidate.first
        except PlaywrightError:
            continue

    return None


def locate_current_query_button(
    page: Page,
) -> Locator:
    """定位第一階段法人查詢表單的查詢按鈕。"""
    exact = page.locator(
        f"xpath={CURRENT_QUERY_BUTTON_XPATH}"
    )

    try:
        exact.first.wait_for(
            state="attached",
            timeout=5000,
        )
        if exact.count() > 0:
            return exact.first
    except PlaywrightError:
        pass

    # 只在第一階段指定的 form 範圍內備援定位，避免誤抓到
    # 原本「法人交通違規繳費紀錄查詢」頁面的查詢按鈕。
    scoped = page.locator(
        "xpath=/html/body/table/tbody/tr[2]/td[1]/div[3]/"
        "div/div[2]/form//a[contains(normalize-space(.), '查詢')]"
    )

    try:
        if scoped.count() > 0:
            return scoped.first
    except PlaywrightError:
        pass

    raise RuntimeError(
        "找不到交通違規（含強制險）法人查詢按鈕；"
        f"XPath：{CURRENT_QUERY_BUTTON_XPATH}"
    )


def open_current_query_corporate_form(
    page: Page,
    entry_url: str,
    timeout_ms: int,
) -> None:
    """
    開啟「交通違規（含強制險）查詢及繳納」法人表單。

    Browser／Context／Page 已關閉時立即向上拋出，交由 main 重建工作
    階段；不可繼續在同一個死亡 page 上輪流嘗試兩個 URL。
    """
    urls: list[str] = []

    for candidate_url in (
        entry_url,
        CURRENT_QUERY_DIRECT_FALLBACK_URL,
    ):
        if candidate_url and candidate_url not in urls:
            urls.append(candidate_url)

    errors: list[str] = []

    for candidate_url in urls:
        try:
            if not page_is_open(page):
                raise RuntimeError(
                    "Target page, context or browser has been closed"
                )

            page.goto(
                candidate_url,
                wait_until="domcontentloaded",
                timeout=timeout_ms,
            )

            try:
                page.wait_for_load_state(
                    "networkidle",
                    timeout=min(timeout_ms, 12000),
                )
            except PlaywrightTimeoutError:
                pass

            try:
                locate_current_query_button(page)
                locate_company_id_input(page)
                return
            except RuntimeError:
                pass

            corporate_tab = (
                locate_current_query_corporate_tab(
                    page
                )
            )

            if corporate_tab is None:
                raise RuntimeError(
                    "找不到交通違規（含強制險）查詢頁的法人頁籤"
                )

            method = click_locator_with_fallback(
                page=page,
                locator=corporate_tab,
                exact_xpath=(
                    CURRENT_QUERY_CORPORATE_TAB_XPATH
                ),
                timeout_ms=timeout_ms,
                label="法人頁籤",
            )

            print(
                f"  第一階段法人頁籤點擊方式：{method}",
                flush=True,
            )

            locate_current_query_button(page)
            locate_company_id_input(page)
            return

        except Exception as exc:
            if is_browser_session_closed_exception(exc):
                raise

            errors.append(
                f"{candidate_url}："
                f"{type(exc).__name__}: {exc}"
            )

    raise RuntimeError(
        "無法開啟交通違規（含強制險）法人查詢表單；"
        + " | ".join(errors)
    )

def click_current_query_button(
    page: Page,
    timeout_ms: int,
) -> str:
    button = locate_current_query_button(page)
    return click_locator_with_fallback(
        page=page,
        locator=button,
        exact_xpath=CURRENT_QUERY_BUTTON_XPATH,
        timeout_ms=timeout_ms,
        label="交通違規（含強制險）查詢按鈕",
    )


def table_rows_from_locator(
    table_locator: Locator,
) -> list[list[dict[str, Any]]]:
    """讀取目標 table 的直屬列，並保留控制項屬性供唯一鍵使用。"""
    return table_locator.evaluate(
        """
        table => {
          const norm = value => (value || '')
            .replace(/\\s+/g, ' ')
            .trim();

          const rows = [
            ...table.querySelectorAll(
              ':scope > thead > tr, '
              + ':scope > tbody > tr, '
              + ':scope > tfoot > tr, '
              + ':scope > tr'
            )
          ];

          return rows.map(row => [
            ...row.children
          ]
            .filter(cell => (
              cell.tagName === 'TH'
              || cell.tagName === 'TD'
            ))
            .map(cell => {
              const controls = [
                ...cell.querySelectorAll(
                  'input, a, button, select, option'
                )
              ].map(control => [
                control.tagName || '',
                control.getAttribute('type') || '',
                control.getAttribute('name') || '',
                control.id || '',
                control.getAttribute('value') || '',
                control.getAttribute('href') || '',
                control.getAttribute('onclick') || '',
                control.getAttribute('data-id') || '',
                control.getAttribute('data-key') || ''
              ].join('|'));

              return {
                text: norm(cell.innerText),
                identity: norm([
                  cell.innerText || '',
                  cell.getAttribute('id') || '',
                  cell.getAttribute('data-id') || '',
                  cell.getAttribute('data-key') || '',
                  ...controls
                ].join(' || ')),
                header: cell.tagName === 'TH'
              };
            }))
            .filter(row => row.length > 0);
        }
        """
    )

def add_current_category_to_row(
    row: list[Any],
    category: str,
) -> list[Any]:
    canonical = list(row[:7])
    canonical += [None] * (7 - len(canonical))

    existing_status = normalize_text(
        canonical[0]
    )

    canonical[0] = (
        f"{category}／{existing_status}"
        if existing_status
        else category
    )

    return canonical


def extract_current_table_at_xpath(
    page: Page,
    category: str,
) -> ExtractedTable | None:
    """擷取目前結果頁籤的精確表格並轉成 155598 的 A:G 結構。"""
    locator = page.locator(
        f"xpath={CURRENT_RESULT_TABLE_XPATH}"
    )

    try:
        if locator.count() <= 0:
            return None

        locator.first.wait_for(
            state="visible",
            timeout=5000,
        )

        payload = table_rows_from_locator(
            locator.first
        )
    except PlaywrightError:
        return None

    plain_rows = [
        [
            normalize_text(
                cell.get("text")
            )
            for cell in row
        ]
        for row in payload
    ]

    combined_text = " ".join(
        " ".join(row)
        for row in plain_rows
    )

    if contains_any(
        combined_text,
        NO_DATA_PATTERNS,
    ):
        return None

    identified = identify_result_header(
        plain_rows
    )

    canonical_rows: list[list[Any]] = []
    canonical_keys: list[str] = []
    headers: list[str] = []

    if identified is not None:
        kind, header_row, columns = identified
        headers = plain_rows[header_row]

        for row_offset, row in enumerate(
            plain_rows[header_row + 1:]
        ):
            canonical = (
                canonicalize_unpaid_row(
                    row,
                    columns,
                )
                if kind == TABLE_KIND_UNPAID
                else None
            )

            if canonical is None:
                continue

            canonical_rows.append(
                add_current_category_to_row(
                    canonical,
                    category,
                )
            )
            raw_identity_row = [
                normalize_text(
                    cell.get("identity")
                    or cell.get("text")
                )
                for cell in payload[
                    header_row + 1 + row_offset
                ]
            ]
            canonical_keys.append(
                record_fingerprint(
                    raw_identity_row
                )
            )
    else:
        # 網站若只更換表頭名稱，仍保留每一列的原始文字，避免整張表遺失。
        data_start = 0
        for index, raw_row in enumerate(payload):
            if any(
                bool(cell.get("header"))
                for cell in raw_row
            ):
                headers = plain_rows[index]
                data_start = index + 1
                break

        for row_offset, row in enumerate(
            plain_rows[data_start:]
        ):
            if not any(row):
                continue

            row_text = " ".join(row)
            if contains_any(
                row_text,
                NO_DATA_PATTERNS,
            ):
                continue

            raw_values: list[Any] = list(row[:7])
            raw_values += [None] * (
                7 - len(raw_values)
            )

            raw_values[0] = (
                f"{category}／{normalize_text(raw_values[0])}"
                if normalize_text(raw_values[0])
                else category
            )

            canonical_rows.append(
                raw_values
            )
            raw_identity_row = [
                normalize_text(
                    cell.get("identity")
                    or cell.get("text")
                )
                for cell in payload[
                    data_start + row_offset
                ]
            ]
            canonical_keys.append(
                record_fingerprint(
                    raw_identity_row
                )
            )

    if not canonical_rows:
        return None

    return ExtractedTable(
        kind=TABLE_KIND_UNPAID,
        title=category,
        headers=headers,
        rows=canonical_rows,
        row_keys=canonical_keys,
    )


def current_result_table_text(
    page: Page,
) -> str:
    locator = page.locator(
        f"xpath={CURRENT_RESULT_TABLE_XPATH}"
    )

    try:
        if locator.count() <= 0:
            return ""
        return normalize_text(
            locator.first.inner_text(
                timeout=3000
            )
        )
    except PlaywrightError:
        return ""


def wait_for_current_table_change(
    page: Page,
    before_text: str,
    timeout_ms: int,
) -> bool:
    deadline = time.monotonic() + min(
        max(timeout_ms, 1000),
        12000,
    ) / 1000.0

    while time.monotonic() < deadline:
        now = current_result_table_text(page)
        body = body_text(page)

        if (
            now != before_text
            or "查無不可線上繳納" in body
            or "無不可線上繳納" in body
        ):
            return True

        time.sleep(0.25)

    return False


def collect_current_violation_tables(
    page: Page,
    timeout_ms: int,
) -> tuple[list[ExtractedTable], list[PageAudit]]:
    """
    逐頁收集第一階段的兩個頁籤。

    先完整爬完「可線上繳納」，確認頁數、總筆數及重複資料均正確後，
    再切換到「不可線上繳納」並執行相同驗證。
    """
    all_tables: list[ExtractedTable] = []
    all_audits: list[PageAudit] = []

    online_tables, online_audits = (
        capture_paginated_group(
            page=page,
            banner_xpath=(
                CURRENT_PAGINATION_BANNER_XPATH
            ),
            section=QUERY_SECTION_CURRENT,
            category=CURRENT_ONLINE_CATEGORY,
            extractor=lambda: (
                [table]
                if (
                    table := extract_current_table_at_xpath(
                        page,
                        CURRENT_ONLINE_CATEGORY,
                    )
                )
                is not None
                else []
            ),
            timeout_ms=timeout_ms,
        )
    )
    all_tables.extend(online_tables)
    all_audits.extend(online_audits)

    before_table_text = current_result_table_text(
        page
    )
    second_tab = page.locator(
        f"xpath={CURRENT_RESULT_SECOND_TAB_XPATH}"
    )

    try:
        second_tab_exists = second_tab.count() > 0
    except PlaywrightError:
        second_tab_exists = False

    if not second_tab_exists:
        raise DataIntegrityError(
            "找不到『不可線上繳納』頁籤，無法完成第一階段完整性檢查",
            {
                "xpath": CURRENT_RESULT_SECOND_TAB_XPATH,
                "result_url": page.url,
            },
        )

    method = click_locator_with_fallback(
        page=page,
        locator=second_tab.first,
        exact_xpath=CURRENT_RESULT_SECOND_TAB_XPATH,
        timeout_ms=timeout_ms,
        label="不可線上繳納頁籤",
    )

    print(
        f"  第一階段第二頁籤點擊方式：{method}",
        flush=True,
    )

    wait_for_current_table_change(
        page,
        before_table_text,
        timeout_ms,
    )

    offline_tables, offline_audits = (
        capture_paginated_group(
            page=page,
            banner_xpath=(
                CURRENT_PAGINATION_BANNER_XPATH
            ),
            section=QUERY_SECTION_CURRENT,
            category=CURRENT_OFFLINE_CATEGORY,
            extractor=lambda: (
                [table]
                if (
                    table := extract_current_table_at_xpath(
                        page,
                        CURRENT_OFFLINE_CATEGORY,
                    )
                )
                is not None
                else []
            ),
            timeout_ms=timeout_ms,
        )
    )
    all_tables.extend(offline_tables)
    all_audits.extend(offline_audits)

    validate_page_audits(all_audits)

    return all_tables, all_audits

def perform_current_violation_query(
    page: Page,
    company: CompanyRow,
    entry_url: str,
    timeout_ms: int,
    max_captcha_attempts: int,
    debug_dir: Path,
    captcha_dir: Path,
    captcha_image_format: str,
) -> QueryOutcome:
    """第一階段：逐頁查詢尚未結案的交通違規及強制險違規。"""
    last_retry_message = ""

    for attempt in range(
        1,
        max_captcha_attempts + 1,
    ):
        try:
            open_current_query_corporate_form(
                page=page,
                entry_url=entry_url,
                timeout_ms=timeout_ms,
            )

            company_input = locate_company_id_input(
                page
            )
            company_input.fill(company.query_id)
            before = body_text(page)

            try:
                captcha_code = acquire_captcha(
                    page=page,
                    company=company,
                    captcha_dir=captcha_dir,
                    captcha_image_format=(
                        captcha_image_format
                    ),
                )
            except Exception as exc:
                if not is_captcha_related_exception(exc):
                    raise

                last_retry_message = (
                    "第一階段 CAPTCHA 辨識未取得四碼："
                    f"{type(exc).__name__}: {exc}"
                )
                print(
                    "  第一階段 CAPTCHA 辨識失敗，將重新查詢"
                    f"（{attempt}/{max_captcha_attempts}）。",
                    flush=True,
                )
                continue

            locate_captcha_input(page).fill(
                captcha_code
            )
            submit_method = click_current_query_button(
                page,
                timeout_ms,
            )
            print(
                f"  第一階段查詢按鈕點擊方式：{submit_method}",
                flush=True,
            )

            text = wait_for_query_result_page(
                page,
                before,
                timeout_ms,
            )

            if contains_any(
                text,
                CAPTCHA_ERROR_PATTERNS,
            ):
                last_retry_message = (
                    "第一階段網站回報驗證碼輸入錯誤"
                )
                continue

            if contains_any(
                text,
                FORM_ERROR_PATTERNS,
            ):
                message = next(
                    pattern
                    for pattern in FORM_ERROR_PATTERNS
                    if pattern in text
                )

                if message in TRANSIENT_FORM_ERROR_PATTERNS:
                    last_retry_message = (
                        f"第一階段網站暫時性回應：{message}"
                    )
                    continue

                return QueryOutcome(
                    STATUS_ERROR,
                    0,
                    [],
                    f"交通違規（含強制險）：{message}",
                    page.url,
                )

            tables, page_audits = (
                collect_current_violation_tables(
                    page,
                    timeout_ms,
                )
            )
            duplicate_keys = validate_page_audits(
                page_audits
            )
            count = sum(
                table.record_count
                for table in tables
            )
            online_count = sum(
                table.record_count
                for table in tables
                if table.category
                == CURRENT_ONLINE_CATEGORY
            )
            offline_count = sum(
                table.record_count
                for table in tables
                if table.category
                == CURRENT_OFFLINE_CATEGORY
            )
            summary = build_pagination_summary(
                page_audits
            )

            return QueryOutcome(
                status=STATUS_SUCCESS,
                record_count=count,
                tables=tables,
                message=(
                    "交通違規（含強制險）查詢完成："
                    f"可線上繳納 {online_count} 筆；"
                    f"不可線上繳納 {offline_count} 筆；"
                    f"{PAGINATION_VERIFIED_MARKER}；"
                    f"{summary}"
                ),
                result_url=page.url,
                page_audits=page_audits,
                duplicate_keys=duplicate_keys,
            )

        except Exception as exc:
            if is_browser_session_closed_exception(exc):
                raise

            if (
                is_captcha_related_exception(exc)
                or is_transient_page_exception(exc)
            ):
                last_retry_message = (
                    "第一階段可重試錯誤："
                    f"{type(exc).__name__}: {exc}"
                )
                continue

            # DataIntegrityError 必須交給 main 立即重爬同一家公司，
            # 不可在此吞掉後處理下一家公司。
            raise

    return QueryOutcome(
        STATUS_RETRY,
        0,
        [],
        (
            last_retry_message
            or (
                "交通違規（含強制險）查詢連續無法完成 "
                f"{max_captcha_attempts} 次"
            )
        ),
        page.url,
        retry_later=True,
    )

def combine_query_outcomes(
    current_outcome: QueryOutcome,
    paid_outcome: QueryOutcome,
) -> QueryOutcome:
    """合併兩階段結果，並再次驗證所有分頁與總筆數。"""
    completed_statuses = {
        STATUS_SUCCESS,
        STATUS_NO_DATA,
    }
    combined_audits = (
        list(current_outcome.page_audits)
        + list(paid_outcome.page_audits)
    )

    if (
        current_outcome.status == STATUS_ERROR
        or paid_outcome.status == STATUS_ERROR
    ):
        return QueryOutcome(
            status=STATUS_ERROR,
            record_count=0,
            tables=[],
            message=(
                "雙階段查詢失敗；"
                f"第一階段：{current_outcome.message}；"
                f"第二階段：{paid_outcome.message}"
            ),
            result_url=(
                current_outcome.result_url
                or paid_outcome.result_url
            ),
            page_audits=combined_audits,
        )

    if (
        current_outcome.status not in completed_statuses
        or paid_outcome.status not in completed_statuses
    ):
        return QueryOutcome(
            status=STATUS_RETRY,
            record_count=0,
            tables=[],
            message=(
                "雙階段查詢尚未全部完成；"
                f"第一階段：{current_outcome.message}；"
                f"第二階段：{paid_outcome.message}"
            ),
            result_url=(
                current_outcome.result_url
                or paid_outcome.result_url
            ),
            retry_later=True,
            page_audits=combined_audits,
        )

    duplicate_keys = validate_page_audits(
        combined_audits
    )
    tables = (
        list(current_outcome.tables)
        + list(paid_outcome.tables)
    )
    total_count = (
        current_outcome.record_count
        + paid_outcome.record_count
    )
    declared_total = sum(
        int(
            page_audit_group_stats(items)[
                "declared_total_records"
            ]
        )
        for items in group_page_audits(
            combined_audits
        ).values()
    )

    if total_count != declared_total:
        raise DataIntegrityError(
            "雙階段合計筆數與三組網站宣告總筆數不一致",
            {
                "first_stage_count": (
                    current_outcome.record_count
                ),
                "second_stage_count": (
                    paid_outcome.record_count
                ),
                "combined_count": total_count,
                "declared_total": declared_total,
                "page_audits": [
                    asdict(item)
                    for item in combined_audits
                ],
            },
        )

    summary = build_pagination_summary(
        combined_audits
    )

    return QueryOutcome(
        status=STATUS_SUCCESS,
        record_count=total_count,
        tables=tables,
        message=(
            f"{COMBINED_QUERY_MESSAGE_MARKER}；"
            f"第一階段 {current_outcome.record_count} 筆；"
            f"第二階段 {paid_outcome.record_count} 筆；"
            f"合計 {total_count} 筆；"
            f"{PAGINATION_VERIFIED_MARKER}；"
            f"{summary}"
        ),
        result_url=(
            current_outcome.result_url
            or paid_outcome.result_url
        ),
        page_audits=combined_audits,
        duplicate_keys=duplicate_keys,
    )

def perform_paid_record_query(
    page: Page,
    company: CompanyRow,
    query_url: str,
    timeout_ms: int,
    max_captcha_attempts: int,
    debug_dir: Path,
    captcha_dir: Path,
    captcha_image_format: str,
    headless: bool,
) -> QueryOutcome:
    """第二階段：逐頁查詢法人交通違規繳納記錄。"""
    last_retry_message = ""

    for attempt in range(
        1,
        max_captcha_attempts + 1,
    ):
        try:
            page.goto(
                query_url,
                wait_until="domcontentloaded",
                timeout=timeout_ms,
            )

            try:
                page.wait_for_load_state(
                    "networkidle",
                    timeout=min(timeout_ms, 15000),
                )
            except PlaywrightTimeoutError:
                pass

            company_input = (
                ensure_query_form_after_optional_login(
                    page,
                    query_url,
                    timeout_ms,
                    headless,
                )
            )
            company_input.fill(company.query_id)
            before = body_text(page)

            try:
                captcha_code = acquire_captcha(
                    page=page,
                    company=company,
                    captcha_dir=captcha_dir,
                    captcha_image_format=(
                        captcha_image_format
                    ),
                )
            except Exception as exc:
                if not is_captcha_related_exception(exc):
                    raise

                last_retry_message = (
                    "CAPTCHA 辨識未取得四碼："
                    f"{type(exc).__name__}: {exc}"
                )
                print(
                    "  CAPTCHA OCR 結果不符合四碼格式，"
                    "將換一張並重新查詢"
                    f"（{attempt}/{max_captcha_attempts}）。",
                    flush=True,
                )
                continue

            locate_captcha_input(page).fill(
                captcha_code
            )
            submit_method = click_query_button(
                page=page,
                timeout_ms=timeout_ms,
            )
            print(
                f"  查詢按鈕點擊方式：{submit_method}",
                flush=True,
            )

            text = wait_for_query_result_page(
                page,
                before,
                timeout_ms,
            )

            if contains_any(
                text,
                CAPTCHA_ERROR_PATTERNS,
            ):
                last_retry_message = (
                    "網站回報驗證碼輸入錯誤"
                )
                print(
                    "  網站回報驗證碼錯誤，"
                    "將重新取得圖片並再次辨識"
                    f"（{attempt}/{max_captcha_attempts}）。",
                    flush=True,
                )
                continue

            if contains_any(
                text,
                FORM_ERROR_PATTERNS,
            ):
                message = next(
                    pattern
                    for pattern in FORM_ERROR_PATTERNS
                    if pattern in text
                )

                if message in TRANSIENT_FORM_ERROR_PATTERNS:
                    last_retry_message = (
                        f"網站暫時性回應：{message}"
                    )
                    continue

                save_debug_artifacts(
                    page,
                    debug_dir,
                    company.query_id,
                    "form_error",
                )
                return QueryOutcome(
                    STATUS_ERROR,
                    0,
                    [],
                    message,
                    page.url,
                )

            try:
                tables, page_audits = (
                    collect_all_result_pages(
                        page,
                        timeout_ms=timeout_ms,
                    )
                )
            except Exception as exc:
                if is_browser_session_closed_exception(exc):
                    raise

                if not is_transient_page_exception(exc):
                    raise

                last_retry_message = (
                    "結果頁導向尚未完成："
                    f"{type(exc).__name__}: {exc}"
                )
                continue

            text = body_text(page)

            if contains_any(
                text,
                CAPTCHA_ERROR_PATTERNS,
            ):
                last_retry_message = (
                    "網站回報驗證碼輸入錯誤"
                )
                continue

            duplicate_keys = validate_page_audits(
                page_audits
            )
            count = sum(
                table.record_count
                for table in tables
            )
            summary = build_pagination_summary(
                page_audits
            )

            return QueryOutcome(
                status=STATUS_SUCCESS,
                record_count=count,
                tables=tables,
                message=(
                    f"繳納紀錄 {count} 筆；"
                    f"{PAGINATION_VERIFIED_MARKER}；"
                    f"{summary}"
                ),
                result_url=page.url,
                page_audits=page_audits,
                duplicate_keys=duplicate_keys,
            )

        except Exception as exc:
            if is_browser_session_closed_exception(exc):
                raise

            if (
                is_captcha_related_exception(exc)
                or is_transient_page_exception(exc)
            ):
                last_retry_message = (
                    f"{type(exc).__name__}: {exc}"
                )
                print(
                    "  遇到可重試的暫時錯誤，將重新查詢"
                    f"（{attempt}/{max_captcha_attempts}）："
                    f"{last_retry_message}",
                    flush=True,
                )
                continue

            raise

    return QueryOutcome(
        STATUS_RETRY,
        0,
        [],
        (
            last_retry_message
            or (
                "CAPTCHA／結果頁連續無法完成 "
                f"{max_captcha_attempts} 次"
            )
        ),
        page.url,
        retry_later=True,
    )

def should_skip_row(
    workbook: Any,
    worksheet: Any,
    company: CompanyRow,
    tracking_columns: dict[str, int],
    resume: bool,
) -> bool:
    """
    只有新版分頁稽核已通過的資料才能續跑略過。

    舊版雖然顯示「成功」，但沒有逐頁驗證與 showbanner 總筆數，
    因此一律重新查詢，避免像 45893188 只留下 23 筆。
    """
    if not resume:
        return False

    status = normalize_text(
        worksheet.cell(
            company.excel_row,
            tracking_columns["違規查詢狀態"],
        ).value
    )

    if status != STATUS_SUCCESS:
        return False

    query_message = normalize_text(
        worksheet.cell(
            company.excel_row,
            tracking_columns["違規查詢訊息"],
        ).value
    )
    integrity_value = normalize_text(
        worksheet.cell(
            company.excel_row,
            tracking_columns["分頁完整性檢查"],
        ).value
    )
    duplicate_value = normalize_text(
        worksheet.cell(
            company.excel_row,
            tracking_columns["重複資料檢查"],
        ).value
    )

    if (
        COMBINED_QUERY_MESSAGE_MARKER
        not in query_message
        or PAGINATION_VERIFIED_MARKER
        not in query_message
        or PAGINATION_VERIFIED_MARKER
        not in integrity_value
        or "無重複" not in duplicate_value
    ):
        return False

    for header in (
        "可線上繳納分頁資訊",
        "不可線上繳納分頁資訊",
        "繳納紀錄分頁資訊",
    ):
        value = normalize_text(
            worksheet.cell(
                company.excel_row,
                tracking_columns[header],
            ).value
        )
        if (
            "頁，共" not in value
            or "已擷取" not in value
            or "重複 0 筆" not in value
        ):
            return False

    record_count_value = worksheet.cell(
        company.excel_row,
        tracking_columns["交通違規筆數"],
    ).value

    try:
        record_count = int(record_count_value)
    except (TypeError, ValueError):
        return False

    if record_count == 0:
        return True

    expected_title = detail_sheet_name(
        company.raw_id,
        company.query_id,
    )

    if expected_title not in workbook.sheetnames:
        return False

    detail_sheet = workbook[expected_title]

    if detail_sheet.max_column < DETAIL_AUDIT_LAST_COLUMN:
        return False

    nonempty_data_rows = 0

    for row in range(1, detail_sheet.max_row + 1):
        if any(
            normalize_text(
                detail_sheet.cell(row, column).value
            )
            for column in range(1, 8)
        ):
            nonempty_data_rows += 1

        if not normalize_text(
            detail_sheet.cell(row, 10).value
        ).startswith("擷取頁碼："):
            return False

        if not normalize_text(
            detail_sheet.cell(row, 13).value
        ).startswith("分頁資訊："):
            return False

        if not normalize_text(
            detail_sheet.cell(row, 14).value
        ).startswith("資料唯一鍵："):
            return False

    return nonempty_data_rows == record_count

def update_main_row(
    worksheet: Any,
    company: CompanyRow,
    outcome: QueryOutcome,
    queried_at: datetime,
    tracking_columns: dict[str, int],
) -> None:
    values = {
        "交通違規筆數": (
            outcome.record_count
            if outcome.status in {
                STATUS_SUCCESS,
                STATUS_NO_DATA,
            }
            else None
        ),
        "違規查詢狀態": outcome.status,
        "最後查詢時間": queried_at,
        "違規查詢訊息": outcome.message,
        "可線上繳納分頁資訊": outcome_group_summary(
            outcome,
            QUERY_SECTION_CURRENT,
            CURRENT_ONLINE_CATEGORY,
        ),
        "不可線上繳納分頁資訊": outcome_group_summary(
            outcome,
            QUERY_SECTION_CURRENT,
            CURRENT_OFFLINE_CATEGORY,
        ),
        "繳納紀錄分頁資訊": outcome_group_summary(
            outcome,
            QUERY_SECTION_PAID,
            PAID_RECORD_CATEGORY,
        ),
        "分頁完整性檢查": outcome_integrity_text(
            outcome
        ),
        "重複資料檢查": outcome_duplicate_text(
            outcome
        ),
    }

    style_source = worksheet.cell(
        company.excel_row,
        tracking_columns["交通違規筆數"],
    )

    for header, value in values.items():
        cell = worksheet.cell(
            company.excel_row,
            tracking_columns[header],
            value,
        )

        if header != "交通違規筆數":
            cell.fill = copy(style_source.fill)
            cell.font = copy(style_source.font)

        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )
        cell.border = Border(
            left=THIN_GRAY,
            right=THIN_GRAY,
            top=THIN_GRAY,
            bottom=THIN_GRAY,
        )

    worksheet.cell(
        company.excel_row,
        tracking_columns["最後查詢時間"],
    ).number_format = "yyyy-mm-dd hh:mm:ss"

def launch_browser(
    playwright: Any,
    headless: bool,
) -> Browser:
    errors: list[Exception] = []

    try:
        return playwright.chromium.launch(
            headless=headless
        )
    except PlaywrightError as exc:
        errors.append(exc)

    try:
        return playwright.chromium.launch(
            channel="chrome",
            headless=headless,
        )
    except PlaywrightError as exc:
        errors.append(exc)

    for executable in (
        shutil.which("chromium"),
        shutil.which("chromium-browser"),
        shutil.which("google-chrome"),
        shutil.which("google-chrome-stable"),
    ):
        if not executable:
            continue

        try:
            return playwright.chromium.launch(
                executable_path=executable,
                headless=headless,
                args=["--no-sandbox"],
            )
        except PlaywrightError as exc:
            errors.append(exc)

    raise RuntimeError(
        "無法啟動瀏覽器。請執行："
        "python -m playwright install chromium"
    ) from (
        errors[-1]
        if errors
        else None
    )


def create_browser_context_and_page(
    browser: Browser,
    timeout_ms: int,
) -> tuple[Any, Page]:
    context = browser.new_context(
        locale="zh-TW",
        viewport={
            "width": 1440,
            "height": 1000,
        },
        ignore_https_errors=False,
    )
    page = context.new_page()
    page.set_default_timeout(timeout_ms)
    return context, page


def safe_close_page(
    page: Page | None,
) -> None:
    if page is None:
        return
    try:
        if not page.is_closed():
            page.close()
    except Exception:
        pass


def safe_close_context(
    context: Any | None,
) -> None:
    if context is None:
        return
    try:
        context.close()
    except Exception:
        pass


def safe_close_browser(
    browser: Browser | None,
) -> None:
    if browser is None:
        return
    try:
        browser.close()
    except Exception:
        pass


def restart_browser_session(
    playwright: Any,
    browser: Browser | None,
    context: Any | None,
    page: Page | None,
    *,
    headless: bool,
    timeout_ms: int,
    reason: str,
) -> tuple[Browser, Any, Page]:
    """關閉失效的 Page／Context，必要時重啟 Browser，再建立新 Page。"""
    print(
        "  偵測到瀏覽器工作階段失效，正在自動重建："
        f"{reason}",
        flush=True,
    )

    safe_close_page(page)
    safe_close_context(context)

    if not browser_is_connected(browser):
        safe_close_browser(browser)
        browser = launch_browser(
            playwright,
            headless,
        )

    try:
        new_context, new_page = (
            create_browser_context_and_page(
                browser,
                timeout_ms,
            )
        )
    except Exception:
        safe_close_browser(browser)
        browser = launch_browser(
            playwright,
            headless,
        )
        new_context, new_page = (
            create_browser_context_and_page(
                browser,
                timeout_ms,
            )
        )

    return browser, new_context, new_page


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "依 Excel 統一編號先查交通違規（含強制險），"
            "再查法人交通違規繳納記錄；"
            "兩階段均逐頁擷取，並以 showbanner 的總頁數／總筆數"
            "驗證完整性與重複資料。"
        )
    )

    parser.add_argument(
        "input",
        nargs="?",
        type=Path,
        default=DEFAULT_INPUT_PATH,
        help=(
            "來源 Excel；預設為專案根目錄下的 "
            r"Data\高雄市.xlsx"
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=(
            "輸出 Excel；預設輸出到 results 資料夾，檔名為 "
            "<來源檔名>_違規明細查詢結果.xlsx"
        ),
    )
    parser.add_argument(
        "--detail-template-sheet",
        default=DEFAULT_DETAIL_TEMPLATE_SHEET,
        help="來源 Excel 的明細格式範本工作表；預設 155598",
    )
    parser.add_argument(
        "--current-query-url",
        default=DEFAULT_CURRENT_QUERY_URL,
        help=(
            "第一階段交通違規（含強制險）入口網址；"
            "找不到時自動改用 penaltyQueryPay"
        ),
    )
    parser.add_argument(
        "--url",
        default=DEFAULT_URL,
        help="第二階段交通違規繳納記錄查詢網址",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=None,
        help="從指定 Excel 列開始",
    )
    parser.add_argument(
        "--end-row",
        type=int,
        default=None,
        help="查到指定 Excel 列為止",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="最多處理幾筆，適合先小量測試",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Playwright 使用無頭模式",
    )
    parser.add_argument(
        "--resume",
        action=argparse.BooleanOptionalAction,
        default=True,
        help=(
            "輸出檔存在時續跑；只有已通過新版分頁完整性"
            "與重複資料檢查的列才會略過"
        ),
    )
    parser.add_argument(
        "--overwrite-output",
        action="store_true",
        help="若輸出檔已存在，刪除並從來源檔重新開始",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30000,
        help="一般網頁逾時毫秒",
    )
    parser.add_argument(
        "--max-captcha-attempts",
        type=int,
        default=5,
        help="單輪處理同一筆資料時，最多重新取得 CAPTCHA 的次數",
    )
    parser.add_argument(
        "--max-captcha-requeues",
        type=int,
        default=1,
        help=(
            "耗盡 CAPTCHA 嘗試後，最多移到工作佇列尾端重做幾次；"
            "預設 1 次"
        ),
    )
    parser.add_argument(
        "--max-integrity-attempts",
        type=int,
        default=3,
        help=(
            "分頁頁數、總筆數或重複資料驗證失敗時，"
            "立即重爬同一家公司最多幾次；預設 3 次"
        ),
    )
    parser.add_argument(
        "--max-browser-restarts",
        type=int,
        default=3,
        help=(
            "Page／BrowserContext／Browser 意外關閉時，"
            "自動重建瀏覽器並立即重跑目前公司最多幾次；預設 3 次"
        ),
    )
    parser.add_argument(
        "--stop-on-integrity-error",
        action=argparse.BooleanOptionalAction,
        default=True,
        help=(
            "完整性驗證連續失敗達上限後停止整批，不處理下一家公司；"
            "預設啟用"
        ),
    )
    parser.add_argument(
        "--delay-min",
        type=float,
        default=1.5,
        help="每筆查詢後最短等待秒數",
    )
    parser.add_argument(
        "--delay-max",
        type=float,
        default=3.0,
        help="每筆查詢後最長等待秒數",
    )
    parser.add_argument(
        "--captcha-dir",
        type=Path,
        default=DEFAULT_CAPTCHA_DIR,
        help="保存每次驗證碼圖片的資料夾",
    )
    parser.add_argument(
        "--captcha-image-format",
        choices=("png", "jpg", "jpeg"),
        default="png",
        help="驗證碼圖片格式；預設 png",
    )
    parser.add_argument(
        "--debug-dir",
        type=Path,
        default=DEFAULT_DEBUG_DIR,
        help="保存除錯 HTML 與截圖的資料夾",
    )
    parser.add_argument(
        "--error-log-dir",
        type=Path,
        default=DEFAULT_ERROR_LOG_DIR,
        help=(
            "分頁不完整、重複資料、未解決錯誤等 JSON 紀錄資料夾；"
            r"預設 results\mvdis_errorLog"
        ),
    )

    return parser

def main(
    argv: Sequence[str] | None = None,
) -> int:
    args = build_parser().parse_args(argv)

    input_path = args.input.expanduser().resolve()
    results_dir = DEFAULT_RESULTS_DIR.expanduser().resolve()

    if not input_path.exists():
        raise SystemExit(f"找不到來源檔：{input_path}")

    if input_path.suffix.lower() != ".xlsx":
        raise SystemExit("目前只支援 .xlsx")

    if (
        args.delay_min < 0
        or args.delay_max < args.delay_min
    ):
        raise SystemExit(
            "delay-min／delay-max 設定錯誤"
        )

    if args.max_captcha_attempts <= 0:
        raise SystemExit(
            "max-captcha-attempts 必須大於 0"
        )

    if args.max_captcha_requeues < 0:
        raise SystemExit(
            "max-captcha-requeues 不可小於 0"
        )

    if args.max_integrity_attempts <= 0:
        raise SystemExit(
            "max-integrity-attempts 必須大於 0"
        )

    if args.max_browser_restarts < 0:
        raise SystemExit(
            "max-browser-restarts 不可小於 0"
        )

    try:
        normalize_captcha_image_format(
            args.captcha_image_format
        )
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc

    if not callable(ocrImage):
        raise SystemExit(
            "from englishAlphanumericOcrApi import ocrImage "
            "匯入的物件不是可呼叫函式"
        )

    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else (
            results_dir
            / (
                f"{input_path.stem}_"
                "違規明細查詢結果.xlsx"
            )
        )
    )

    if output_path == input_path:
        raise SystemExit(
            "輸出檔不能與來源檔相同；本程式不會覆寫原始 Excel"
        )

    new_output = (
        not output_path.exists()
        or args.overwrite_output
    )

    if args.overwrite_output and output_path.exists():
        output_path.unlink()

    if new_output:
        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )
        shutil.copy2(input_path, output_path)
        print(f"已建立新輸出檔：{output_path}")
    else:
        print(
            "偵測到既有輸出檔，將續跑："
            f"{output_path}"
        )

    workbook = load_workbook(output_path)
    (
        main_sheet,
        header_start_row,
        data_start_row,
        id_col,
        name_col,
    ) = find_main_sheet_and_columns(workbook)

    detail_template_sheet_name = (
        ensure_detail_template(
            workbook,
            main_sheet.title,
            args.detail_template_sheet,
            input_path,
        )
    )

    if new_output:
        clear_stale_numeric_detail_sheets(
            workbook,
            main_sheet.title,
        )

    tracking_columns = ensure_tracking_columns(
        main_sheet,
        header_start_row,
    )

    if new_output:
        reset_tracking_values(
            main_sheet,
            data_start_row,
            tracking_columns,
        )

    log_sheet = ensure_log_sheet(workbook)
    pagination_audit_sheet = (
        ensure_pagination_audit_sheet(workbook)
    )
    duplicate_audit_sheet = (
        ensure_duplicate_audit_sheet(workbook)
    )

    companies = iter_companies(
        main_sheet,
        data_start_row,
        id_col,
        name_col,
        args.start_row,
        args.end_row,
        args.limit,
    )

    if not companies:
        raise SystemExit(
            "指定範圍內沒有可用的統一編號／登記編號"
        )

    captcha_dir = args.captcha_dir.expanduser().resolve()
    debug_dir = args.debug_dir.expanduser().resolve()
    error_log_dir = (
        args.error_log_dir.expanduser().resolve()
    )
    error_log_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    print(
        f"主工作表：{main_sheet.title}；"
        f"ID 欄：{get_column_letter(id_col)}；"
        f"名稱欄：{get_column_letter(name_col)}；"
        f"待掃描：{len(companies)} 筆。"
    )
    print(f"來源 Excel：{input_path}")
    print(f"輸出 Excel：{output_path}")
    print(
        "第二階段分頁資訊 XPath："
        f"{PAID_PAGINATION_BANNER_XPATH}"
    )
    print(
        "第一階段分頁資訊 XPath："
        f"{CURRENT_PAGINATION_BANNER_XPATH}"
    )
    print(
        "第一階段結果表格 XPath："
        f"{CURRENT_RESULT_TABLE_XPATH}"
    )
    print(
        "第一階段第二頁籤 XPath："
        f"{CURRENT_RESULT_SECOND_TAB_XPATH}"
    )
    print(
        "結果 JSON XPath（僅作目前頁備援）："
        f"{RESULT_JSON_XPATH}"
    )
    print(
        "違規明細格式範本："
        f"{args.detail_template_sheet} "
        f"→ {detail_template_sheet_name}"
    )
    print(f"驗證碼圖片資料夾：{captcha_dir}")
    print(f"錯誤 JSON 資料夾：{error_log_dir}")
    print(
        "完整性失敗策略：同一家公司立即重爬 "
        f"{args.max_integrity_attempts} 次；"
        f"達上限後"
        f"{'停止整批' if args.stop_on_integrity_error else '記錄錯誤後繼續'}。"
    )
    print(
        "瀏覽器失效策略：自動重建 Page／Context／Browser，"
        f"同一家公司立即重跑最多 {args.max_browser_restarts} 次。"
    )

    processed = 0
    success = 0
    zero_result_success = 0
    no_data = 0
    errors = 0
    pending = 0
    skipped = 0
    deferred = 0
    queue_runs = 0
    integrity_retries = 0
    browser_restarts = 0

    work_queue = deque(
        (company, 0)
        for company in companies
    )

    with sync_playwright() as playwright:
        browser = launch_browser(
            playwright,
            args.headless,
        )
        context, page = create_browser_context_and_page(
            browser,
            args.timeout,
        )

        try:
            while work_queue:
                company, requeue_count = (
                    work_queue.popleft()
                )

                if (
                    not browser_is_connected(browser)
                    or not page_is_open(page)
                ):
                    browser, context, page = restart_browser_session(
                        playwright,
                        browser,
                        context,
                        page,
                        headless=args.headless,
                        timeout_ms=args.timeout,
                        reason="開始處理公司前檢查到工作階段已失效",
                    )
                    browser_restarts += 1

                if should_skip_row(
                    workbook,
                    main_sheet,
                    company,
                    tracking_columns,
                    args.resume,
                ):
                    skipped += 1
                    print(
                        f"[已完成 {processed}/{len(companies)}] "
                        f"{company.query_id} {company.name}："
                        "已通過分頁完整性與重複檢查，略過。"
                    )
                    continue

                queue_runs += 1
                retry_label = (
                    ""
                    if requeue_count == 0
                    else (
                        "；隊尾重試 "
                        f"{requeue_count}/"
                        f"{args.max_captcha_requeues}"
                    )
                )
                print(
                    f"[執行 {queue_runs}；"
                    f"已完成 {processed}/{len(companies)}；"
                    f"佇列尚有 {len(work_queue)}] "
                    f"查詢 {company.query_id} {company.name}"
                    f"{retry_label}",
                    flush=True,
                )

                queried_at = datetime.now()
                outcome: QueryOutcome | None = None
                stop_after_current = False

                integrity_attempt = 1
                company_browser_restarts = 0

                while (
                    integrity_attempt
                    <= args.max_integrity_attempts
                ):
                    try:
                        print(
                            "  [階段 1/2] 交通違規（含強制險）查詢及繳納",
                            flush=True,
                        )
                        current_outcome = (
                            perform_current_violation_query(
                                page=page,
                                company=company,
                                entry_url=(
                                    args.current_query_url
                                ),
                                timeout_ms=args.timeout,
                                max_captcha_attempts=(
                                    args.max_captcha_attempts
                                ),
                                debug_dir=debug_dir,
                                captcha_dir=captcha_dir,
                                captcha_image_format=(
                                    args.captcha_image_format
                                ),
                            )
                        )

                        print(
                            "  [階段 2/2] 法人交通違規繳納記錄查詢",
                            flush=True,
                        )
                        paid_outcome = (
                            perform_paid_record_query(
                                page=page,
                                company=company,
                                query_url=args.url,
                                timeout_ms=args.timeout,
                                max_captcha_attempts=(
                                    args.max_captcha_attempts
                                ),
                                debug_dir=debug_dir,
                                captcha_dir=captcha_dir,
                                captcha_image_format=(
                                    args.captcha_image_format
                                ),
                                headless=args.headless,
                            )
                        )

                        outcome = combine_query_outcomes(
                            current_outcome,
                            paid_outcome,
                        )
                        break

                    except KeyboardInterrupt:
                        print(
                            "\n使用者中止。已保存目前進度。"
                        )
                        atomic_save(
                            workbook,
                            output_path,
                        )
                        return 130

                    except DataIntegrityError as exc:
                        integrity_retries += 1
                        save_debug_artifacts(
                            page,
                            debug_dir,
                            company.query_id,
                            (
                                "integrity_error_"
                                f"{integrity_attempt}"
                            ),
                        )
                        error_path = write_error_json(
                            error_log_dir,
                            company,
                            phase=(
                                "分頁完整性驗證失敗_"
                                f"第{integrity_attempt}次"
                            ),
                            message=str(exc),
                            page=(
                                page
                                if page_is_open(page)
                                else None
                            ),
                            exception=exc,
                            details={
                                "integrity_attempt": (
                                    integrity_attempt
                                ),
                                "max_integrity_attempts": (
                                    args.max_integrity_attempts
                                ),
                            },
                        )
                        print(
                            "  → 分頁完整性未通過；"
                            f"{exc}；錯誤 JSON：{error_path}",
                            flush=True,
                        )

                        if (
                            integrity_attempt
                            < args.max_integrity_attempts
                        ):
                            integrity_attempt += 1
                            print(
                                "  不處理下一家公司，立即重新爬取"
                                f"目前公司（{integrity_attempt}/"
                                f"{args.max_integrity_attempts}）。",
                                flush=True,
                            )
                            time.sleep(
                                random.uniform(
                                    args.delay_min,
                                    args.delay_max,
                                )
                            )
                            continue

                        outcome = QueryOutcome(
                            status=STATUS_ERROR,
                            record_count=0,
                            tables=[],
                            message=(
                                "分頁完整性驗證連續失敗 "
                                f"{args.max_integrity_attempts} 次；"
                                f"{exc}；已輸出 JSON，不會把不完整資料"
                                "標成成功"
                            ),
                            result_url=safe_page_url(page),
                        )
                        stop_after_current = (
                            args.stop_on_integrity_error
                        )
                        break

                    except Exception as exc:
                        if is_browser_session_closed_exception(exc):
                            company_browser_restarts += 1
                            browser_restarts += 1

                            error_path = write_error_json(
                                error_log_dir,
                                company,
                                phase=(
                                    "瀏覽器工作階段失效_"
                                    f"第{company_browser_restarts}次"
                                ),
                                message=(
                                    f"{type(exc).__name__}: {exc}"
                                ),
                                page=None,
                                exception=exc,
                                details={
                                    "browser_restart_attempt": (
                                        company_browser_restarts
                                    ),
                                    "max_browser_restarts": (
                                        args.max_browser_restarts
                                    ),
                                },
                            )

                            if (
                                company_browser_restarts
                                <= args.max_browser_restarts
                            ):
                                browser, context, page = (
                                    restart_browser_session(
                                        playwright,
                                        browser,
                                        context,
                                        page,
                                        headless=args.headless,
                                        timeout_ms=args.timeout,
                                        reason=(
                                            f"{type(exc).__name__}: {exc}"
                                        ),
                                    )
                                )
                                print(
                                    "  已建立新的 Page／Context，"
                                    "立即重跑目前公司，不移到下一筆；"
                                    f"錯誤 JSON：{error_path}",
                                    flush=True,
                                )
                                continue

                            outcome = QueryOutcome(
                                status=STATUS_RETRY,
                                record_count=0,
                                tables=[],
                                message=(
                                    "瀏覽器工作階段連續失效 "
                                    f"{company_browser_restarts} 次；"
                                    f"{type(exc).__name__}: {exc}"
                                ),
                                result_url="",
                                retry_later=True,
                            )
                            print(
                                "  瀏覽器自動重建已達上限；"
                                f"錯誤 JSON：{error_path}",
                                flush=True,
                            )
                            break

                        save_debug_artifacts(
                            page,
                            debug_dir,
                            company.query_id,
                            "exception",
                        )
                        retryable_exception = (
                            is_captcha_related_exception(exc)
                            or is_transient_page_exception(exc)
                        )
                        outcome = QueryOutcome(
                            status=(
                                STATUS_RETRY
                                if retryable_exception
                                else STATUS_ERROR
                            ),
                            record_count=0,
                            tables=[],
                            message=(
                                f"{type(exc).__name__}: {exc}"
                            ),
                            result_url=safe_page_url(page),
                            retry_later=retryable_exception,
                        )
                        error_path = write_error_json(
                            error_log_dir,
                            company,
                            phase="查詢例外",
                            message=outcome.message,
                            page=(
                                page
                                if page_is_open(page)
                                else None
                            ),
                            exception=exc,
                            outcome=outcome,
                        )
                        print(
                            f"  例外已寫入 JSON：{error_path}",
                            flush=True,
                        )
                        break

                if outcome is None:
                    raise RuntimeError(
                        "內部錯誤：查詢結束後沒有 QueryOutcome"
                    )

                if outcome.retry_later:
                    if (
                        requeue_count
                        < args.max_captcha_requeues
                    ):
                        next_requeue_count = (
                            requeue_count + 1
                        )
                        deferred_outcome = QueryOutcome(
                            status=STATUS_RETRY,
                            record_count=0,
                            tables=[],
                            message=(
                                f"{outcome.message}；"
                                "已移到工作佇列尾端，其他資料完成後"
                                "將自動重做"
                                f"（隊尾重試 {next_requeue_count}/"
                                f"{args.max_captcha_requeues}）"
                            ),
                            result_url=outcome.result_url,
                            retry_later=True,
                            page_audits=(
                                outcome.page_audits
                            ),
                            duplicate_keys=(
                                outcome.duplicate_keys
                            ),
                        )
                        remove_existing_detail_variants(
                            workbook,
                            company.raw_id,
                            company.query_id,
                        )
                        update_main_row(
                            main_sheet,
                            company,
                            deferred_outcome,
                            queried_at,
                            tracking_columns,
                        )
                        append_log(
                            log_sheet,
                            company,
                            deferred_outcome,
                            queried_at,
                        )
                        replace_company_audit_rows(
                            pagination_audit_sheet,
                            duplicate_audit_sheet,
                            company,
                            deferred_outcome,
                            queried_at,
                        )
                        work_queue.append(
                            (company, next_requeue_count)
                        )
                        error_path = write_error_json(
                            error_log_dir,
                            company,
                            phase="待重試_已移到隊尾",
                            message=deferred_outcome.message,
                            page=(page if page_is_open(page) else None),
                            outcome=deferred_outcome,
                        )
                        atomic_save(
                            workbook,
                            output_path,
                        )
                        deferred += 1
                        print(
                            "  → 待重試；"
                            f"{company.query_id} 已排到佇列尾端；"
                            f"JSON：{error_path}",
                            flush=True,
                        )

                        if work_queue:
                            time.sleep(
                                random.uniform(
                                    args.delay_min,
                                    args.delay_max,
                                )
                            )
                        continue

                    outcome = QueryOutcome(
                        status=STATUS_RETRY,
                        record_count=0,
                        tables=[],
                        message=(
                            f"{outcome.message}；"
                            "已達本次執行允許的隊尾重試上限 "
                            f"{args.max_captcha_requeues} 次；"
                            "保留為待重試，下次使用 --resume 執行時"
                            "會自動重做"
                        ),
                        result_url=outcome.result_url,
                        page_audits=outcome.page_audits,
                        duplicate_keys=outcome.duplicate_keys,
                    )

                update_main_row(
                    main_sheet,
                    company,
                    outcome,
                    queried_at,
                    tracking_columns,
                )

                if outcome.status == STATUS_SUCCESS:
                    validate_page_audits(
                        outcome.page_audits
                    )
                    if outcome.record_count > 0:
                        write_detail_sheet(
                            workbook,
                            detail_template_sheet_name,
                            company,
                            outcome,
                            queried_at,
                        )
                    else:
                        remove_existing_detail_variants(
                            workbook,
                            company.raw_id,
                            company.query_id,
                        )
                        zero_result_success += 1
                    success += 1

                elif outcome.status == STATUS_NO_DATA:
                    remove_existing_detail_variants(
                        workbook,
                        company.raw_id,
                        company.query_id,
                    )
                    no_data += 1

                elif outcome.status == STATUS_RETRY:
                    # 不保留舊版可能漏頁的明細，避免使用者誤把舊資料當成本次結果。
                    remove_existing_detail_variants(
                        workbook,
                        company.raw_id,
                        company.query_id,
                    )
                    pending += 1

                else:
                    # 完整性或不可恢復錯誤時，移除舊的不完整明細工作表。
                    remove_existing_detail_variants(
                        workbook,
                        company.raw_id,
                        company.query_id,
                    )
                    errors += 1

                replace_company_audit_rows(
                    pagination_audit_sheet,
                    duplicate_audit_sheet,
                    company,
                    outcome,
                    queried_at,
                )
                append_log(
                    log_sheet,
                    company,
                    outcome,
                    queried_at,
                )

                if outcome.status != STATUS_SUCCESS:
                    error_path = write_error_json(
                        error_log_dir,
                        company,
                        phase=(
                            "最終錯誤"
                            if outcome.status == STATUS_ERROR
                            else "最終待重試"
                        ),
                        message=outcome.message,
                        page=(page if page_is_open(page) else None),
                        outcome=outcome,
                    )
                    print(
                        f"  問題已寫入 JSON：{error_path}",
                        flush=True,
                    )

                atomic_save(workbook, output_path)
                processed += 1

                print(
                    f"  → {outcome.status}；"
                    f"違規筆數={outcome.record_count}；"
                    f"{outcome.message}"
                )

                if stop_after_current:
                    print(
                        "完整性驗證已達重試上限。"
                        "依預設設定停止整批，沒有處理下一家公司。"
                    )
                    print(f"輸出檔：{output_path}")
                    print(f"錯誤 JSON：{error_log_dir}")
                    return 3

                if work_queue:
                    time.sleep(
                        random.uniform(
                            args.delay_min,
                            args.delay_max,
                        )
                    )

        finally:
            safe_close_page(page)
            safe_close_context(context)
            safe_close_browser(browser)

    print(
        f"完成。最終處理 {processed} 筆："
        f"成功 {success}（其中零筆 {zero_result_success}）、"
        f"舊版無資料 {no_data}、"
        f"待重試 {pending}、"
        f"錯誤 {errors}；"
        f"略過 {skipped}、"
        f"曾移至隊尾 {deferred} 次、"
        f"完整性立即重爬 {integrity_retries} 次、"
        f"瀏覽器自動重建 {browser_restarts} 次。"
    )
    print(f"輸出檔：{output_path}")
    print(f"錯誤 JSON：{error_log_dir}")

    return 0 if errors == 0 else 2



if __name__ == "__main__":
    raise SystemExit(main())
